import io
import json

from openpyxl import load_workbook

from packages.core.services.document_service import (
    SPREADSHEET_CHARTS_SHEET,
    SPREADSHEET_EDITOR_PAYLOAD_PREFIX,
    _spreadsheet_content_to_xlsx_bytes,
)


def test_spreadsheet_editor_payload_generates_xlsx_charts_and_metadata():
    payload = {
        "data": [
            ["Month", "Sales", "Cost", "Zip", "Profit"],
            ["Jan", "10", "7", "00123", "=B2-C2"],
            ["Feb", "20", "11", "00456", "=B3-C3"],
            ["Mar", "35", "19", "00789", "=B4-C4"],
        ],
        "charts": [
            {
                "id": "bar1",
                "type": "bar",
                "title": "Sales Bar",
                "labelColumn": 0,
                "valueColumn": 1,
                "startRow": 1,
                "endRow": 3,
            },
            {
                "id": "line1",
                "type": "line",
                "title": "Cost Line",
                "labelColumn": 0,
                "valueColumn": 2,
                "startRow": 1,
                "endRow": 3,
            },
            {
                "id": "pie1",
                "type": "pie",
                "title": "Sales Pie",
                "labelColumn": 0,
                "valueColumn": 1,
                "startRow": 1,
                "endRow": 3,
            },
            {
                "id": "profit1",
                "type": "bar",
                "title": "Profit Bar",
                "labelColumn": 0,
                "valueColumn": 4,
                "startRow": 1,
                "endRow": 3,
            },
        ],
    }

    content = SPREADSHEET_EDITOR_PAYLOAD_PREFIX + json.dumps(payload)
    wb = load_workbook(io.BytesIO(_spreadsheet_content_to_xlsx_bytes(content)))
    ws = wb.active

    assert len(ws._charts) == 4
    assert [chart.__class__.__name__ for chart in ws._charts] == ["BarChart", "LineChart", "PieChart", "BarChart"]
    assert ws["B2"].value == 10
    assert ws["C2"].value == 7
    assert ws["D2"].value == "00123"
    assert ws["E2"].value == "=B2-C2"
    assert ws["E2"].data_type == "f"
    assert SPREADSHEET_CHARTS_SHEET in wb.sheetnames
    assert wb[SPREADSHEET_CHARTS_SHEET].sheet_state == "hidden"
    assert json.loads(wb[SPREADSHEET_CHARTS_SHEET]["A1"].value)["charts"][0]["title"] == "Sales Bar"


def test_plain_csv_generates_xlsx_without_chart_metadata():
    wb = load_workbook(
        io.BytesIO(_spreadsheet_content_to_xlsx_bytes("Name,Zip,Formula\nAlice,00123,=1+2\nBob,00456,=2+3"))
    )
    ws = wb.active

    assert len(ws._charts) == 0
    assert SPREADSHEET_CHARTS_SHEET not in wb.sheetnames
    assert ws["A2"].value == "Alice"
    assert ws["B2"].value == "00123"
    assert ws["C2"].value == "=1+2"
    assert ws["C2"].data_type == "f"
