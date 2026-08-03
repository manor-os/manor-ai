"""
Workflow Runner — executes step-based workflows to completion or pause.

Step types supported:
  - agent:     Run agentic_loop with a skill/prompt and tools
  - tool:      Execute a single tool call directly
  - condition: Evaluate expression against workflow variables, branch
  - wait:      Pause execution (HITL approval, timer, external event)
  - parallel:  Run multiple sub-steps concurrently

Usage:
    from packages.core.ai.workflow_runner import WorkflowRunner
    await WorkflowRunner().run(workflow_run_id)
"""
from __future__ import annotations

import ast
import asyncio
import base64
import hashlib
import html
import json
import logging
import operator
import re
import time
from copy import deepcopy
from datetime import datetime, timedelta, timezone
from email.utils import parsedate_to_datetime
from types import SimpleNamespace
from typing import Any
from xml.etree import ElementTree

from jsonschema import Draft202012Validator

from packages.core.ai.runtime import (
    ChatSurface,
    runtime_attach_and_persist_workflow_runner_result,
    runtime_execute_workflow_agent_loop,
    runtime_execute_workflow_tool_step,
    runtime_invoke_skill,
    runtime_merge_prompt_appendix,
    runtime_prepare_named_tool_surface_for_turn,
    runtime_prepare_prompt_appendix_for_turn,
    runtime_prepare_trace_envelope_for_turn,
    runtime_prompt_with_output_schema,
    runtime_request_for_surface_turn,
    runtime_workflow_run_context,
    runtime_workflow_tool_context_args,
)
from packages.core.database import async_session
from packages.core.models.base import generate_ulid
from packages.core.models.workflow import WorkflowDefinition, WorkflowRun
from packages.core.services.workflow_run_trace import (
    DEFINITION_CHANGED_ERROR,
    append_execution_trace,
    build_definition_snapshot,
    summarize_trace_text,
    workflow_definition_changed,
    workflow_definition_fingerprint,
)

logger = logging.getLogger(__name__)

# ── Constants ────────────────────────────────────────────────────────────────
_STEP_TIMEOUT_SECS = 300.0
_DURABLE_ORCHESTRATION_NODE_TYPES = {"stage", "subworkflow", "foreach_subworkflow"}
_AGENT_MAX_ROUNDS = 20
_AGENT_TEMPERATURE = 0.7
# Timer waits up to this run inline (slept); longer ones pause for an external
# resume. Kept under the ~100s origin timeout so an inline run never 524s.
_INLINE_WAIT_CAP_SECS = 90.0

_OPS = {
    "==": operator.eq,
    "!=": operator.ne,
    ">=": operator.ge,
    "<=": operator.le,
    ">": operator.gt,
    "<": operator.lt,
}

# Node types whose output is a pure, reproducible function of their inputs and
# whose only side effect is via ``output_var`` / ``next_override`` (re-applied
# from the cached result) — so reusing a cached result is safe. External /
# non-deterministic types (llm, http, connector, rag, tool, agent, notify) are
# never cached by default. ``transform`` mutates run.variables directly (not via
# output_var), so it is intentionally excluded.
_CACHEABLE_TYPES = frozenset({"condition", "switch", "merge"})

# Max nesting for sub-workflow (Execute Workflow) calls — a runaway-recursion
# backstop; depth is carried in the child run's trigger_data.
_MAX_SUBWORKFLOW_DEPTH = 5


class _StageRunProxy:
    """Mutable operation-local view of a parent run."""

    def __init__(
        self,
        parent: WorkflowRun,
        *,
        variables: dict[str, Any],
        operation_results: dict[str, dict[str, Any]],
    ) -> None:
        for field in (
            "id",
            "workflow_id",
            "entity_id",
            "workspace_id",
            "started_by",
            "trigger_source",
            "attempt_number",
            "retry_of_run_id",
            "retry_from_step_id",
            "lineage_root_run_id",
            "lineage_is_legacy",
            "started_at",
        ):
            setattr(self, field, getattr(parent, field, None))
        self.variables = variables
        self.step_results = {
            **dict(parent.step_results or {}),
            **operation_results,
        }
        self.trigger_data = deepcopy(parent.trigger_data or {})
        self.current_step_id: str | None = None
        self.status = "running"
        self.error: str | None = None
        self.completed_at = None
        self.effective_attempt_number = parent.effective_attempt_number


def _stage_operation_checkpoint(result: dict[str, Any]) -> dict[str, Any]:
    """Keep only the control metadata required to resume an internal graph."""
    retained_fields = (
        "status",
        "continued",
        "skipped",
        "next_override",
        "condition_result",
        "wait_type",
        "duration_seconds",
        "resume_at",
        "auto_resume_scheduled",
        "resumed",
        "resumed_at",
        "completed_at",
        "decision",
        "approved",
        "approved_by",
        "approved_at",
        "attempts",
        "step_id",
        "fingerprint",
    )
    return {
        field: deepcopy(result[field])
        for field in retained_fields
        if field in result
    }


def _workflow_sandbox_url() -> str:
    """Resolve the sandbox URL at call time so tests/config reloads stay reliable."""
    try:
        from packages.core.config import get_settings

        return str(get_settings().SANDBOX_SERVICE_URL or "").strip()
    except Exception:
        return ""


def _is_cacheable(step: dict) -> bool:
    """Whether a step's result may be reused from cache (ComfyUI-style).

    ``config.cache_policy`` overrides the default: ``never`` opts out, ``cache``
    /``auto`` opts in (e.g. to memoise an idempotent llm/http call on re-run).
    """
    policy = (step.get("config") or {}).get("cache_policy")
    if policy == "never":
        return False
    if policy in ("cache", "auto"):
        return True
    return step.get("type") in _CACHEABLE_TYPES


def _step_fingerprint(step: dict, variables: dict) -> str:
    """Stable hash of a step's identity + inputs (ComfyUI IS_CHANGED analogue).

    Conservative: fingerprints over the full variable set, so any upstream
    change invalidates downstream cache (never stale, may under-reuse). A future
    refinement could fingerprint only the variables a step actually references.
    """
    payload = {
        "type": step.get("type"),
        "config": step.get("config", {}),
        "next": step.get("next"),
        "true_next": step.get("true_next"),
        "false_next": step.get("false_next"),
        "vars": variables,
    }
    blob = json.dumps(payload, sort_keys=True, default=str)
    return hashlib.sha256(blob.encode("utf-8")).hexdigest()


def _utc_now() -> datetime:
    return datetime.now(timezone.utc)


def workflow_approval_decision_metadata(
    config: dict[str, Any],
    *,
    decision: Any,
    actor_id: str,
    decided_at: datetime | None = None,
) -> dict[str, Any]:
    """Validate one configured approval decision and build durable metadata."""
    normalized_decision = str(decision or "").strip()
    normalized_actor = str(actor_id or "").strip()
    if not normalized_decision or not normalized_actor:
        raise ValueError("Workflow approval requires a decision and actor")
    options = {
        str(option).strip()
        for option in (config.get("options") or [])
        if str(option or "").strip()
    }
    if options and normalized_decision not in options:
        raise ValueError("Unsupported Workflow approval choice")
    approval_values = {
        str(value).strip().lower()
        for value in (
            config.get("approval_values")
            or ["approve", "approved", "accept", "accepted", "yes"]
        )
    }
    return {
        "decision": normalized_decision,
        "approved": normalized_decision.lower() in approval_values,
        "approved_by": normalized_actor,
        "approved_at": (decided_at or _utc_now()).isoformat(),
    }


def workflow_stage_wait_context(
    run: WorkflowRun,
    current_step: dict[str, Any] | None,
) -> tuple[str, dict[str, Any], dict[str, Any]] | None:
    """Resolve the paused internal wait and its durable stage state."""
    if not isinstance(current_step, dict) or current_step.get("type") != "stage":
        return None
    stage_id = str(current_step.get("id") or "")
    stage_execution = (
        (run.variables or {}).get("__stage_execution")
        if isinstance((run.variables or {}).get("__stage_execution"), dict)
        else {}
    )
    stage_state = stage_execution.get(stage_id)
    if not isinstance(stage_state, dict):
        return None
    operation_id = str(
        stage_state.get("paused_operation_id")
        or stage_state.get("current_operation_id")
        or ""
    ).strip()
    operations = (
        (current_step.get("config") or {}).get("operations")
        if isinstance(current_step.get("config"), dict)
        else []
    )
    operation = next(
        (
            candidate
            for candidate in operations or []
            if isinstance(candidate, dict)
            and str(candidate.get("id") or "") == operation_id
            and candidate.get("type") == "wait"
        ),
        None,
    )
    if not operation_id or operation is None:
        return None
    return operation_id, operation, deepcopy(stage_state)


def complete_workflow_stage_wait(
    run: WorkflowRun,
    current_step: dict[str, Any],
    context: tuple[str, dict[str, Any], dict[str, Any]],
    *,
    metadata: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Complete one paused internal wait and make its parent stage runnable."""
    operation_id, operation, stage_state = context
    stage_id = str(current_step.get("id") or "")
    operation_results = deepcopy(stage_state.get("operation_results") or {})
    previous = dict(operation_results.get(operation_id) or {})
    completed_at = _utc_now().isoformat()
    previous.update({
        "status": "completed",
        "resumed": True,
        "resumed_at": completed_at,
        "completed_at": completed_at,
        **(metadata or {}),
    })
    operation_results[operation_id] = _stage_operation_checkpoint(previous)
    stage_state.update({
        "status": "running",
        "current_operation_id": None,
        "pending_operation_ids": [],
        "operation_results": operation_results,
        "paused_operation_id": None,
        "failed_operation_id": None,
    })
    updated_vars = dict(run.variables or {})
    stage_execution = deepcopy(updated_vars.get("__stage_execution") or {})
    stage_execution[stage_id] = stage_state
    updated_vars["__stage_execution"] = stage_execution
    run.variables = updated_vars
    step_results = dict(run.step_results or {})
    step_results.pop(stage_id, None)
    run.step_results = step_results
    operation_config = (
        operation.get("config")
        if isinstance(operation.get("config"), dict)
        else {}
    )
    append_execution_trace(
        run,
        node={
            **operation,
            "id": f"{stage_id}.{operation_id}",
            "config": {
                **operation_config,
                "chat_projection": "hidden",
            },
        },
        status="completed",
        result=previous,
    )
    return previous


async def _project_workflow_chat_safely(db, projector, **kwargs) -> None:
    """Keep best-effort Chat projection failures out of the Workflow transaction."""
    try:
        async with db.begin_nested():
            await projector(db, **kwargs)
    except Exception:
        logger.debug("Workflow Chat projection skipped", exc_info=True)


def _render_template(template: str, variables: dict) -> str:
    """Replace ``{{var}}`` / ``{{var.field.subfield}}`` placeholders.

    Plain ``{{x}}`` reads ``variables["x"]``; dotted ``{{x.a.b}}`` walks into
    dict values (e.g. a step output stored as a dict). Unknown refs are left
    as-is so a missing variable is visible rather than silently blanked.
    """
    def _replacer(m: re.Match) -> str:
        key = m.group(1).strip()
        val = _lookup_reference(key, variables, missing=m.group(0))
        return str(val) if val is not None else m.group(0)
    # ``[^{}]+`` allows spaces and punctuation inside a ref so imported node
    # names resolve; non-variable expressions (``{{ $env.X }}``) simply fail
    # the lookup above and are left visible, unchanged.
    return re.sub(r"\{\{([^{}]+?)\}\}", _replacer, template)


_SINGLE_REF_RE = re.compile(r"^\s*\{\{([^{}]+?)\}\}\s*$")


def _walk_path(value: Any, parts: list[str], missing: Any = None) -> Any:
    """Walk dict keys and numeric list indexes used by imported expressions."""
    current = value
    for part in parts:
        if isinstance(current, dict) and part in current:
            current = current[part]
        elif isinstance(current, (list, tuple)) and part.isdigit():
            index = int(part)
            if index >= len(current):
                return missing
            current = current[index]
        else:
            return missing
    return current


def _lookup_reference(key: str, variables: dict, missing: Any = None) -> Any:
    """Resolve a reference while allowing n8n node names to contain dots.

    ``{{My node.0.title}}`` first tries the whole key, then the longest variable
    name prefix (``My node``) before walking the remaining dict/list path.
    """
    if key in variables:
        return variables[key]
    candidates = [name for name in variables if key.startswith(f"{name}.")]
    if candidates:
        root = max(candidates, key=len)
        return _walk_path(variables[root], key[len(root) + 1:].split("."), missing=missing)
    parts = key.split(".")
    if parts[0] not in variables:
        return missing
    return _walk_path(variables[parts[0]], parts[1:], missing=missing)


def _resolve_binding(value: Any, variables: dict) -> Any:
    """Resolve an input/output binding value against the variable scope.

    A value that is *exactly* a single reference (``{{step}}`` or
    ``{{step.field}}``) resolves to the raw upstream value, preserving its type
    (dict / list / number) so a whole object can be mapped through. Anything
    else — mixed text, multiple refs, a literal — is string-interpolated via
    :func:`_render_template`. Non-string values pass through unchanged.
    """
    if not isinstance(value, str):
        return value
    m = _SINGLE_REF_RE.match(value)
    if m:
        key = m.group(1).strip()
        return _lookup_reference(key, variables, missing=value)
    return _render_template(value, variables)


def _resolve_structure(value: Any, variables: dict) -> Any:
    """Resolve Workflow references recursively while preserving JSON value types."""
    if isinstance(value, dict):
        return {key: _resolve_structure(item, variables) for key, item in value.items()}
    if isinstance(value, list):
        return [_resolve_structure(item, variables) for item in value]
    return _resolve_binding(value, variables)


def _parse_agent_output(value: Any, output_format: Any) -> Any:
    if str(output_format or "text").strip().lower() != "json":
        return value
    if isinstance(value, dict):
        return value
    from packages.core.services.skill_bundle import extract_json_object

    try:
        return extract_json_object(str(value or ""))
    except ValueError as exc:
        raise ValueError(f"Agent output must be a valid JSON object: {exc}") from exc


def _merge_project_state(base: dict[str, Any], patch: dict[str, Any]) -> dict[str, Any]:
    merged = deepcopy(base)
    for key, value in patch.items():
        if isinstance(value, dict) and isinstance(merged.get(key), dict):
            merged[key] = _merge_project_state(merged[key], value)
        else:
            merged[key] = deepcopy(value)
    return merged


def _project_state_list(state: dict[str, Any], path: str) -> list[Any]:
    parts = [part for part in str(path or "").split(".") if part]
    if not parts:
        raise ValueError("Workflow project list update requires a path")
    current: dict[str, Any] = state
    for part in parts[:-1]:
        child = current.get(part)
        if child is None:
            child = {}
            current[part] = child
        if not isinstance(child, dict):
            raise ValueError(f"Workflow project list path is not an object: {path}")
        current = child
    value = current.get(parts[-1])
    if value is None:
        value = []
        current[parts[-1]] = value
    if not isinstance(value, list):
        raise ValueError(f"Workflow project list path is not an array: {path}")
    return value


def _upsert_project_list_item(items: list[Any], item: Any, key: str) -> None:
    if not isinstance(item, dict):
        raise ValueError("Workflow project list item must be an object")
    item_key = str(item.get(key) or "").strip()
    if not item_key:
        raise ValueError(f"Workflow project list item has no {key}")
    for index, existing in enumerate(items):
        if isinstance(existing, dict) and str(existing.get(key) or "").strip() == item_key:
            items[index] = deepcopy(item)
            return
    items.append(deepcopy(item))


def _reconcile_project_list_items(
    items: list[Any],
    replacements: list[Any],
    key: str,
    ordered_keys: list[Any],
) -> None:
    def keyed(values: list[Any], source: str) -> dict[str, dict[str, Any]]:
        by_key: dict[str, dict[str, Any]] = {}
        for item in values:
            if not isinstance(item, dict):
                raise ValueError(f"Workflow project {source} item must be an object")
            item_key = str(item.get(key) or "").strip()
            if not item_key:
                raise ValueError(f"Workflow project {source} item has no {key}")
            if item_key in by_key:
                raise ValueError(
                    f"Workflow project {source} has duplicate {key}: {item_key}"
                )
            by_key[item_key] = item
        return by_key

    existing_by_key = keyed(items, "list")
    replacements_by_key = keyed(replacements, "reconcile")
    desired_keys: list[str] = []
    for value in ordered_keys:
        item_key = str(value or "").strip()
        if not item_key:
            raise ValueError(f"Workflow project reconcile key {key} cannot be empty")
        if item_key in desired_keys:
            raise ValueError(
                f"Workflow project reconcile has duplicate {key}: {item_key}"
            )
        desired_keys.append(item_key)

    unexpected = set(replacements_by_key) - set(desired_keys)
    if unexpected:
        raise ValueError(
            f"Workflow project reconcile replacement is outside declared {key}s: "
            f"{sorted(unexpected)[0]}"
        )

    reconciled: list[dict[str, Any]] = []
    for item_key in desired_keys:
        item = replacements_by_key.get(item_key) or existing_by_key.get(item_key)
        if item is None:
            raise ValueError(
                f"Workflow project reconcile is missing {key}: {item_key}"
            )
        reconciled.append(deepcopy(item))
    items[:] = reconciled


def _remove_project_list_items(
    items: list[Any],
    field: str,
    values: list[Any],
) -> None:
    if not field:
        raise ValueError("Workflow project list removal requires a field")
    removal_values = {
        str(value).strip()
        for value in values
        if str(value or "").strip()
    }
    for item in items:
        if not isinstance(item, dict):
            raise ValueError("Workflow project list removal item must be an object")
    items[:] = [
        item
        for item in items
        if str(item.get(field) or "").strip() not in removal_values
    ]


def _trusted_workflow_context(run: WorkflowRun) -> dict[str, Any]:
    trigger_data = run.trigger_data if isinstance(run.trigger_data, dict) else {}
    trusted = trigger_data.get("_workflow_runtime_context")
    return deepcopy(trusted) if isinstance(trusted, dict) else {}


def _set_trusted_workflow_context(run: WorkflowRun, **values: Any) -> None:
    trigger_data = dict(run.trigger_data or {})
    trusted = _trusted_workflow_context(run)
    for key, value in values.items():
        if value is None or not str(value).strip():
            trusted.pop(key, None)
        else:
            trusted[key] = value
    trigger_data["_workflow_runtime_context"] = trusted
    run.trigger_data = trigger_data


def _child_workflow_context(run: WorkflowRun, seed: dict[str, Any]) -> dict[str, Any]:
    trusted = _trusted_workflow_context(run)
    for key in (
        "workflow_project_id",
        "workflow_action_grant_id",
        "workflow_scene_id",
        "workflow_batch_capture",
        "approved_plan_version",
    ):
        if key not in seed:
            continue
        value = seed.get(key)
        if value is None or not str(value).strip():
            trusted.pop(key, None)
        else:
            trusted[key] = value
    return trusted


_INPUT_SNAPSHOT_MAX_KEYS = 40
_INPUT_SNAPSHOT_MAX_LEN = 600


def _input_snapshot(step: dict, variables: dict) -> dict:
    """A capped, display-safe view of the data a step sees at run time — the
    variable scope after its named inputs are bound — for the UI's per-node
    Input panel. Values are stringified and truncated so step_results stays small.
    ``variables`` is mutated (binding) — pass a copy.
    """
    _bind_inputs(step.get("config") or {}, variables)
    out: dict = {}
    for i, (k, v) in enumerate(variables.items()):
        if i >= _INPUT_SNAPSHOT_MAX_KEYS:
            out["…"] = f"+{len(variables) - _INPUT_SNAPSHOT_MAX_KEYS} more"
            break
        try:
            s = v if isinstance(v, (int, float, bool)) or v is None else (
                v if isinstance(v, str) else json.dumps(v, ensure_ascii=False, default=str)
            )
        except (TypeError, ValueError):
            s = str(v)
        if isinstance(s, str) and len(s) > _INPUT_SNAPSHOT_MAX_LEN:
            s = s[:_INPUT_SNAPSHOT_MAX_LEN] + "…"
        out[k] = s
    return out


def _schema_validation_failure(
    schema: Any,
    value: Any,
    *,
    code: str,
) -> dict | None:
    if not isinstance(schema, dict):
        return None
    try:
        errors = sorted(
            Draft202012Validator(schema).iter_errors(value),
            key=lambda error: [str(part) for part in error.absolute_path],
        )
    except Exception as exc:
        return {
            "status": "failed",
            "code": code,
            "error": f"Invalid JSON Schema: {exc}",
        }
    if not errors:
        return None
    error = errors[0]
    path = "$"
    for part in error.absolute_path:
        path += f"[{part}]" if isinstance(part, int) else f".{part}"
    return {
        "status": "failed",
        "code": code,
        "error": f"{path}: {error.message}",
    }


def _continues_on_error(step: dict) -> bool:
    """n8n-style "On Error → Continue": a failed step advances the workflow
    instead of halting it. The step is still recorded as failed (no output is
    stored, so downstream ``{{id}}`` refs stay visibly unresolved)."""
    return str((step.get("config") or {}).get("on_error", "stop")).strip().lower() == "continue"


def _final_run_output(steps: list[dict], step_results: dict) -> Any:
    """Pick the durable business result for a completed workflow run.

    Prefer an explicit End node output. Older definitions used End as a label,
    so fall back to the latest completed non-control step in that case.
    """
    by_id = {step.get("id"): step for step in steps}
    for step in reversed(steps):
        if step.get("type") != "end":
            continue
        result = step_results.get(step.get("id")) or {}
        output = result.get("output")
        if output not in (None, "") and output != step.get("name"):
            return output
    for step_id, result in reversed(list((step_results or {}).items())):
        step = by_id.get(step_id) or {}
        if step.get("type") in {"trigger", "webhook", "end", "note"}:
            continue
        if result.get("status") != "completed" or result.get("skipped"):
            continue
        output = result.get("output")
        if output not in (None, ""):
            return output
    return None


def _resolve_list(raw: Any, variables: dict) -> list:
    """Resolve a config value to a list — a ``{{ref}}`` to a list var, an inline
    list, or a single value wrapped. Used by filter / aggregate / loop-style nodes."""
    val = _resolve_binding(raw, variables) if isinstance(raw, str) else raw
    if isinstance(val, (list, tuple)):
        return list(val)
    return [] if val is None else [val]


def _dig(item: Any, field: Any) -> Any:
    """Pull a (possibly dotted) field from a dict item; pass non-dicts through."""
    if not field:
        return item
    cur = item
    for part in str(field).split("."):
        if isinstance(cur, dict):
            cur = cur.get(part)
        else:
            return None
    return cur


def _parse_dt(value: str) -> datetime | None:
    """Best-effort parse of an ISO-ish datetime string (Z suffix tolerated)."""
    s = value.strip()
    if not s or s.lower() == "now":
        return None
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except (ValueError, TypeError):
        try:
            parsed = parsedate_to_datetime(s)
            return parsed.replace(tzinfo=timezone.utc) if parsed.tzinfo is None else parsed
        except (ValueError, TypeError, OverflowError):
            return None


def _plain_feed_text(value: str) -> str:
    return re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", html.unescape(value or ""))).strip()


def _parse_rss_feed(xml_text: str) -> list[dict[str, Any]]:
    """Parse RSS/Atom into the item shape n8n's RSS Read node emits."""
    root = ElementTree.fromstring(xml_text)

    def local_name(tag: str) -> str:
        return tag.rsplit("}", 1)[-1]

    entries = [element for element in root.iter() if local_name(element.tag) in {"item", "entry"}]
    articles: list[dict[str, Any]] = []
    for entry in entries:
        item: dict[str, Any] = {}
        for child in list(entry):
            local = local_name(child.tag)
            text_value = "".join(child.itertext()).strip()
            if local == "link":
                item["link"] = child.attrib.get("href") or text_value
            elif local == "encoded":
                item["content:encoded"] = text_value
            elif local in {"published", "updated", "date"}:
                item.setdefault("pubDate", text_value)
                item.setdefault("isoDate", text_value)
            elif local == "description":
                item["contentSnippet"] = _plain_feed_text(text_value)
                item.setdefault("content", text_value)
            elif local == "summary":
                item["contentSnippet"] = _plain_feed_text(text_value)
            elif local == "content":
                item["content"] = text_value
                item.setdefault("contentSnippet", _plain_feed_text(text_value))
            elif local in {"title", "guid", "id", "creator", "author", "pubDate"}:
                item[local] = text_value
        if item.get("title"):
            # A feed can contain full HTML articles. Keep enough for a useful
            # digest without inflating workflow state and model prompts forever.
            for key in ("content:encoded", "content", "contentSnippet"):
                if isinstance(item.get(key), str):
                    item[key] = item[key][:12000]
            articles.append(item)
    return articles


def _simple_css_xpath(selector: str) -> str:
    """Translate the small, common CSS subset used by n8n HTML templates."""
    paths: list[str] = []
    for group in str(selector or "").split(","):
        tokens = [token for token in group.strip().split() if token]
        if not tokens:
            continue
        path = "."
        for index, token in enumerate(tokens):
            tag_match = re.match(r"^[A-Za-z][A-Za-z0-9_-]*|^\*", token)
            tag = tag_match.group(0) if tag_match else "*"
            predicates: list[str] = []
            id_match = re.search(r"#([A-Za-z0-9_-]+)", token)
            if id_match:
                predicates.append(f"@id={id_match.group(1)!r}")
            for class_name in re.findall(r"\.([A-Za-z0-9_-]+)", token):
                predicates.append(
                    "contains(concat(' ', normalize-space(@class), ' '), "
                    f"{' ' + class_name + ' '!r})"
                )
            predicate = "[" + " and ".join(predicates) + "]" if predicates else ""
            # CSS selection includes the context element itself. lxml's ``.//``
            # only searches descendants, so a common n8n selector such as
            # ``html`` otherwise misses the document root and returns empty.
            axis = "/descendant-or-self::" if index == 0 else "//"
            path += f"{axis}{tag}{predicate}"
        paths.append(path)
    return " | ".join(paths) or ".//*"


def _html_extract_fields(source: Any, fields: list[dict]) -> dict[str, Any]:
    """Apply n8n HTML-node extraction rules to one HTTP/body value."""
    from lxml import html as lxml_html

    if isinstance(source, dict) and "body" in source:
        source = source["body"]
    if isinstance(source, (dict, list)):
        source = json.dumps(source, ensure_ascii=False, default=str)
    document = lxml_html.fromstring(str(source or ""))
    extracted: dict[str, Any] = {}
    for field in fields:
        nodes = document.xpath(_simple_css_xpath(str(field.get("selector") or "")))
        values: list[Any] = []
        for node in nodes:
            clone = deepcopy(node)
            for skip in str(field.get("skip_selectors") or "").split(","):
                skip = skip.strip()
                if not skip:
                    continue
                for child in clone.xpath(_simple_css_xpath(skip)):
                    parent = child.getparent()
                    if parent is not None:
                        parent.remove(child)
            mode = str(field.get("return_value") or "text").lower()
            if mode == "attribute":
                value = node.get(str(field.get("attribute") or ""), "")
            elif mode == "html":
                value = lxml_html.tostring(clone, encoding="unicode")
            else:
                value = " ".join(" ".join(clone.itertext()).split())
            values.append(value)
        extracted[str(field.get("key"))] = values if field.get("return_array") else (values[0] if values else "")
    return extracted


def _coerce_typed(value: Any, type_: Any) -> Any:
    """Coerce a bound value to its declared parameter type — ComfyUI-style
    widget typing (INT/FLOAT/STRING) applied to our named bindings. Best-effort:
    an unparseable value passes through unchanged rather than erroring.
    ``any`` / ``image`` (opaque ref) and unknown types are pass-through.
    """
    if type_ == "number":
        if isinstance(value, bool) or isinstance(value, (int, float)):
            return value
        try:
            f = float(str(value).strip())
            return int(f) if f.is_integer() else f
        except (ValueError, TypeError):
            return value
    if type_ == "json":
        if isinstance(value, (dict, list)):
            return value
        try:
            return json.loads(value) if isinstance(value, str) else value
        except (ValueError, TypeError):
            return value
    if type_ == "text":
        return value if isinstance(value, str) else ("" if value is None else str(value))
    return value


def _bind_inputs(config: dict, variables: dict) -> None:
    """Merge a step's explicit named inputs (``config.inputs``) into ``variables``.

    Each ``{key, value, type?}`` binds ``key`` to the resolved (and type-coerced)
    ``value`` so the step's own config (prompt / body / args) can reference it as
    ``{{key}}``. Mutates ``variables`` in place; later rows can read earlier ones.
    """
    for item in config.get("inputs") or []:
        if not isinstance(item, dict):
            continue
        key = str(item.get("key") or item.get("name") or "").strip()
        if key:
            resolved = _resolve_binding(item.get("value", ""), variables)
            variables[key] = _coerce_typed(resolved, item.get("type"))


def _resolve_value(raw: str, variables: dict) -> Any:
    """Resolve a token from a condition expression.

    Checks workflow variables first, then tries to parse as a Python literal
    (number, string, bool, None).  Falls back to the raw string.
    """
    stripped = raw.strip()

    if not stripped.startswith(("'", '"')):
        missing = object()
        resolved = _lookup_reference(stripped, variables, missing=missing)
        if resolved is not missing:
            return resolved

    # ``true`` / ``false`` convenience
    if stripped.lower() == "true":
        return True
    if stripped.lower() == "false":
        return False

    # Try literal (int, float, string with quotes, None, etc.)
    try:
        return ast.literal_eval(stripped)
    except Exception:
        return stripped


# ── WorkflowRunner ───────────────────────────────────────────────────────────

class WorkflowRunner:
    """Execute a WorkflowRun through its step graph.

    Each invocation of ``run()`` processes steps until the workflow completes,
    pauses (wait step / HITL), or fails.

    Pass ``cache_index`` (built via :meth:`prime_cache_from_results` off a prior
    run) to enable ComfyUI-style incremental re-execution: cacheable steps whose
    fingerprint matches reuse the prior result instead of re-executing.
    """

    def __init__(self, cache_index: dict | None = None) -> None:
        self._cache_index = cache_index or {}

    @staticmethod
    def prime_cache_from_results(step_results: dict | None) -> dict:
        """Build a {fingerprint: result} cache from a prior run's step_results.

        Only completed, cacheable steps that recorded a fingerprint are kept.
        Volatile fields (duration, cached flag) are stripped from the reused
        result so re-application is clean.
        """
        index: dict = {}
        for res in (step_results or {}).values():
            fp = res.get("fingerprint")
            if fp and res.get("status") == "completed":
                cached = {k: v for k, v in res.items() if k not in ("duration_ms", "cached")}
                index[fp] = cached
        return index

    # ── Public entry point ───────────────────────────────────────────────

    async def run(self, workflow_run_id: str, progress=None) -> None:
        """Execute a workflow run to completion or pause.

        ``progress`` is an optional async callback invoked with
        ``{"id": step_id, "status": ...}`` as each step starts ("running") and
        finishes — used by the streaming run endpoint to light up the canvas
        node-by-node. ``None`` (default) leaves behaviour unchanged.
        """
        async with async_session() as db:
            from sqlalchemy import select
            result = await db.execute(
                select(WorkflowRun).where(WorkflowRun.id == workflow_run_id)
            )
            run = result.scalar_one_or_none()
            if not run:
                logger.warning("WorkflowRunner: run %s not found", workflow_run_id)
                return
            if run.status not in ("running", "pending"):
                logger.debug("WorkflowRunner: run %s status=%s, skipping", workflow_run_id, run.status)
                return

            wf_result = await db.execute(
                select(WorkflowDefinition).where(WorkflowDefinition.id == run.workflow_id)
            )
            workflow = wf_result.scalar_one_or_none()
            if not workflow:
                run.status = "failed"
                run.error = "Workflow definition not found"
                from packages.core.ledger.adapters import record_workflow_run_status
                await record_workflow_run_status(db, run)
                await db.commit()
                return
            if workflow_definition_changed(workflow, run):
                run.status = "failed"
                run.error = DEFINITION_CHANGED_ERROR
                run.completed_at = _utc_now()
                from packages.core.services.workflow_chat_projection import (
                    project_workflow_run_status,
                )

                await _project_workflow_chat_safely(
                    db,
                    project_workflow_run_status,
                    run=run,
                )
                await db.commit()
                return

            # Mark running if pending
            if run.status == "pending":
                run.status = "running"
                run.started_at = run.started_at or _utc_now()
                await db.commit()

            try:
                await self._run_loop(workflow, run, db, progress)
            except Exception as exc:
                logger.error("WorkflowRunner: run %s failed: %s", workflow_run_id, exc, exc_info=True)
                await db.rollback()
                run.status = "failed"
                run.error = str(exc)
                run.completed_at = _utc_now()
                from packages.core.services.workflow_chat_projection import (
                    project_workflow_run_status,
                )

                await _project_workflow_chat_safely(
                    db,
                    project_workflow_run_status,
                    run=run,
                )
                await db.commit()

            # Ledger (M1): the loop settled the run (completed/failed/paused);
            # repeat calls dedupe on the per-run+event idempotency key.
            from packages.core.ledger.adapters import record_workflow_run_status
            await record_workflow_run_status(db, run)
            await db.commit()

            # n8n-style Error trigger: a failed run fires every error-handler
            # binding so a workflow can react to failures (alert / log / recover).
            if run.status == "failed":
                await self._dispatch_error_handlers(run, db)
            if run.status in ("completed", "failed"):
                await self._propagate_subworkflow_result(run, db)

    async def _propagate_subworkflow_result(self, child: WorkflowRun, db) -> None:
        """Continue a paused parent after its nested workflow finishes.

        A subworkflow that pauses for approval/timer also pauses its parent. When
        the child later completes (or fails), this atomically replaces the
        parent's paused subworkflow result and resumes the parent graph.
        """
        trigger_data = child.trigger_data or {}
        parent_id = trigger_data.get("parent_run_id")
        parent_step_id = trigger_data.get("parent_step_id")
        if not parent_id or not parent_step_id:
            return
        if trigger_data.get("parent_foreach_item_key"):
            await self._propagate_foreach_subworkflow_result(
                child,
                db,
                parent_id=str(parent_id),
                parent_step_id=str(parent_step_id),
                item_key=str(trigger_data["parent_foreach_item_key"]),
                progress_key=str(trigger_data.get("parent_foreach_progress_key") or ""),
            )
            return

        from sqlalchemy import select

        parent = (await db.execute(
            select(WorkflowRun)
            .where(WorkflowRun.id == parent_id)
            .with_for_update()
        )).scalar_one_or_none()
        if not parent or parent.status != "paused" or parent.current_step_id != parent_step_id:
            return
        previous = dict((parent.step_results or {}).get(parent_step_id) or {})
        if previous.get("subrun_id") != child.id:
            return

        workflow = (await db.execute(
            select(WorkflowDefinition).where(WorkflowDefinition.id == parent.workflow_id)
        )).scalar_one_or_none()
        step = next(
            (s for s in (workflow.steps if workflow else []) if s.get("id") == parent_step_id),
            None,
        )
        if not workflow or not step:
            parent.status = "failed"
            parent.error = "Parent workflow or subworkflow step no longer exists"
            parent.completed_at = _utc_now()
            from packages.core.ledger.adapters import record_workflow_run_status
            await record_workflow_run_status(db, parent)
            await db.commit()
            return
        if workflow_definition_changed(workflow, parent):
            parent.status = "failed"
            parent.error = DEFINITION_CHANGED_ERROR
            parent.completed_at = _utc_now()
            await db.commit()
            return

        if child.status == "failed":
            completed_at = _utc_now().isoformat()
            failed = {
                **previous,
                "status": "failed",
                "error": child.error or "subworkflow failed",
                "subrun_id": child.id,
                "completed_at": completed_at,
            }
            self._record_step_result(step, failed, parent)
            append_execution_trace(
                parent,
                node=step,
                status="failed",
                result=failed,
            )
            parent.status = "failed"
            parent.error = failed["error"]
            parent.completed_at = _utc_now()
            from packages.core.ledger.adapters import record_workflow_run_status
            await record_workflow_run_status(db, parent)
            await db.commit()
            await self._dispatch_error_handlers(parent, db)
            return

        completed_at = _utc_now().isoformat()
        completed = {
            **previous,
            "status": "completed",
            "output": dict(child.variables or {}),
            "subrun_id": child.id,
            "resumed": True,
            "resumed_at": completed_at,
            "completed_at": completed_at,
            "output_var": (step.get("config") or {}).get("output_var"),
        }
        self._record_step_result(step, completed, parent)
        append_execution_trace(
            parent,
            node=step,
            status="completed",
            result=completed,
        )
        parent.status = "running"
        parent.error = None
        parent_run_id = parent.id
        await db.commit()
        await WorkflowRunner(cache_index=self._cache_index).run(parent_run_id)

    async def _propagate_foreach_subworkflow_result(
        self,
        child: WorkflowRun,
        db,
        *,
        parent_id: str,
        parent_step_id: str,
        item_key: str,
        progress_key: str,
    ) -> None:
        """Update one durable foreach item and resume the parent barrier."""
        from sqlalchemy import select

        parent = (await db.execute(
            select(WorkflowRun)
            .where(WorkflowRun.id == parent_id)
            .with_for_update()
        )).scalar_one_or_none()
        if (
            parent is None
            or parent.status != "paused"
            or parent.current_step_id != parent_step_id
            or not progress_key
        ):
            return

        workflow = (await db.execute(
            select(WorkflowDefinition).where(
                WorkflowDefinition.id == parent.workflow_id,
            )
        )).scalar_one_or_none()
        if workflow is None:
            parent.status = "failed"
            parent.error = "Parent workflow no longer exists"
            parent.completed_at = _utc_now()
            await db.commit()
            return
        if workflow_definition_changed(workflow, parent):
            parent.status = "failed"
            parent.error = DEFINITION_CHANGED_ERROR
            parent.completed_at = _utc_now()
            await db.commit()
            return

        variables = dict(parent.variables or {})
        progress = dict(variables.get(progress_key) or {})
        items = [dict(item) for item in progress.get("items") or []]
        matched = False
        for item in items:
            if item.get("key") != item_key or item.get("subrun_id") != child.id:
                continue
            matched = True
            item["status"] = child.status
            item["output"] = dict(child.variables or {}) if child.status == "completed" else None
            item["error"] = child.error if child.status == "failed" else None
            break
        if not matched:
            return

        progress["items"] = items
        variables[progress_key] = progress
        parent.variables = variables

        active = [
            item for item in items
            if item.get("status") in {"running", "paused"}
        ]
        step_results = dict(parent.step_results or {})
        previous = dict(step_results.get(parent_step_id) or {})
        previous["items"] = self._foreach_public_items(items)
        previous["subrun_ids"] = [
            str(item["subrun_id"])
            for item in active
            if item.get("subrun_id")
        ]
        if active:
            step_results[parent_step_id] = previous
            parent.step_results = step_results
            await db.commit()
            return

        step_results.pop(parent_step_id, None)
        parent.step_results = step_results
        parent.status = "running"
        parent.error = None
        parent_run_id = parent.id
        await db.commit()
        await WorkflowRunner(cache_index=self._cache_index).run(parent_run_id)

    async def _dispatch_error_handlers(self, run: WorkflowRun, db) -> None:
        """Start error-handler workflows bound with trigger_type=='error'.

        Skipped for runs that were themselves started by an error trigger, so a
        failing handler can't recurse. Best-effort: dispatch failures never mask
        the original error.
        """
        if (run.trigger_source or "") == "error":
            return
        try:
            from packages.core.services import workflow_service as svc

            handler_runs = await svc.dispatch_trigger(
                db,
                run.entity_id,
                trigger_type="error",
                trigger_data={
                    "failed_workflow_id": run.workflow_id,
                    "failed_run_id": run.id,
                    "error": run.error,
                },
                started_by=run.started_by,
            )
            if handler_runs:
                await db.commit()
                for r in handler_runs:
                    self.enqueue(r.id)
        except Exception:  # noqa: BLE001 — never let error-handling break the run
            logger.warning("error-handler dispatch failed for run %s", run.id, exc_info=True)

    # ── Core loop ────────────────────────────────────────────────────────

    async def _run_loop(
        self, workflow: WorkflowDefinition, run: WorkflowRun, db, progress=None,
    ) -> None:
        """Loop: find runnable steps -> execute -> advance -> repeat."""
        trigger_data = run.trigger_data if isinstance(run.trigger_data, dict) else {}
        entrypoint_context = trigger_data.get("_workspace_chat_entrypoint")
        projects_to_workspace_chat = (
            str(run.trigger_source or "") == "workspace_chat"
            and isinstance(entrypoint_context, dict)
            and bool(entrypoint_context.get("enabled"))
        )

        async def _emit(step, status, result=None):
            step_id = step["id"]
            append_execution_trace(
                run,
                node=step,
                status=status,
                result=result,
            )
            if progress:
                try:
                    await progress({"id": step_id, "status": status})
                except Exception:  # noqa: BLE001 — progress is best-effort telemetry
                    pass
            if not projects_to_workspace_chat:
                return
            try:
                from packages.core.services.workflow_chat_projection import (
                    project_workflow_step,
                )

                await _project_workflow_chat_safely(
                    db,
                    project_workflow_step,
                    run=run,
                    step=step,
                    status=status,
                    result=result,
                )
            except Exception:
                logger.debug("Workflow Chat step projection import skipped", exc_info=True)

        steps = workflow.steps or []
        max_iterations = len(steps) * 3  # safety cap

        for _ in range(max_iterations):
            runnable = self._find_runnable_steps(workflow, run)
            if not runnable:
                # No more steps to run — check if we're done
                if self._all_steps_done(workflow, run):
                    run.status = "completed"
                    run.completed_at = _utc_now()
                    final_output = _final_run_output(steps, run.step_results or {})
                    if final_output is not None:
                        run.variables = {**(run.variables or {}), "__result": final_output}
                else:
                    # Could be waiting (paused) or stuck
                    if run.status != "paused":
                        run.status = "failed"
                        run.error = "No runnable steps and workflow not complete"
                        run.completed_at = _utc_now()
                from packages.core.services.workflow_chat_projection import (
                    project_workflow_run_status,
                )

                await _project_workflow_chat_safely(
                    db,
                    project_workflow_run_status,
                    run=run,
                )
                await db.commit()
                return

            # Parallel steps: execute concurrently
            if len(runnable) > 1:
                for step in runnable:
                    await _emit(step, "running")
                if projects_to_workspace_chat:
                    await db.commit()
                tasks = [
                    self._execute_step_safe(step, run, db)
                    for step in runnable
                ]
                results = await asyncio.gather(*tasks)
                batch = list(zip(runnable, results))
                for step, result in batch:
                    if result.get("status") == "failed" and _continues_on_error(step):
                        result["continued"] = True
                    self._record_step_result(step, result, run)
                ordered_terminals = sorted(
                    enumerate(batch),
                    key=lambda item: (
                        item[1][1].get("completed_at") is None,
                        str(item[1][1].get("completed_at") or ""),
                        item[0],
                    ),
                )
                for _, (step, result) in ordered_terminals:
                    await _emit(
                        step,
                        "skipped" if result.get("skipped") else result.get("status"),
                        result,
                    )
                for step, result in batch:
                    if result.get("status") == "paused":
                        run.status = "paused"
                        from packages.core.services.workflow_chat_projection import project_workflow_run_status

                        await _project_workflow_chat_safely(
                            db,
                            project_workflow_run_status,
                            run=run,
                        )
                        await db.commit()
                        return
                    if result.get("status") == "failed" and not _continues_on_error(step):
                        run.status = "failed"
                        run.error = summarize_trace_text(
                            result.get("error"),
                            fallback=f"Step {step['id']} failed",
                        )
                        run.completed_at = _utc_now()
                        from packages.core.services.workflow_chat_projection import project_workflow_run_status

                        await _project_workflow_chat_safely(
                            db,
                            project_workflow_run_status,
                            run=run,
                        )
                        await db.commit()
                        return
            else:
                step = runnable[0]
                await _emit(step, "running")
                if projects_to_workspace_chat:
                    await db.commit()
                result = await self._execute_step_safe(step, run, db)
                if result.get("status") == "failed" and _continues_on_error(step):
                    result["continued"] = True
                self._record_step_result(step, result, run)
                await _emit(
                    step,
                    "skipped" if result.get("skipped") else result.get("status"),
                    result,
                )
                if result.get("status") == "paused":
                    run.status = "paused"
                    from packages.core.services.workflow_chat_projection import project_workflow_run_status

                    await _project_workflow_chat_safely(
                        db,
                        project_workflow_run_status,
                        run=run,
                    )
                    await db.commit()
                    return
                if result.get("status") == "failed" and not _continues_on_error(step):
                    run.status = "failed"
                    run.error = summarize_trace_text(
                        result.get("error"),
                        fallback=f"Step {step['id']} failed",
                    )
                    run.completed_at = _utc_now()
                    from packages.core.services.workflow_chat_projection import project_workflow_run_status

                    await _project_workflow_chat_safely(
                        db,
                        project_workflow_run_status,
                        run=run,
                    )
                    await db.commit()
                    return

            await db.commit()

        # Exhausted safety cap
        run.status = "failed"
        run.error = "Exceeded maximum iteration limit"
        run.completed_at = _utc_now()
        from packages.core.services.workflow_chat_projection import project_workflow_run_status

        await _project_workflow_chat_safely(
            db,
            project_workflow_run_status,
            run=run,
        )
        await db.commit()

    # ── Step dispatch ────────────────────────────────────────────────────

    async def _execute_step_safe(
        self, step: dict, run: WorkflowRun, db,
    ) -> dict:
        """Execute a step with timeout and error handling."""
        step_id = step["id"]
        run.current_step_id = step_id
        config = step.get("config", {})
        resolved_inputs = dict(run.variables or {})
        _bind_inputs(config, resolved_inputs)
        input_snapshot = _input_snapshot(step, dict(run.variables or {}))
        start = time.monotonic()

        input_failure = _schema_validation_failure(
            config.get("input_schema"),
            resolved_inputs,
            code="input_schema_validation_failed",
        )
        if input_failure is not None:
            input_failure.update({
                "duration_ms": (time.monotonic() - start) * 1000,
                "step_id": step_id,
                "inputs": input_snapshot,
                "completed_at": _utc_now().isoformat(),
            })
            return input_failure

        # ComfyUI-style incremental re-execution: cacheable steps carry a
        # fingerprint of their inputs; a matching fingerprint in the cache reuses
        # the prior result instead of re-executing.
        fingerprint = _step_fingerprint(step, dict(run.variables or {})) if _is_cacheable(step) else None
        if fingerprint is not None:
            cached = self._cache_index.get(fingerprint)
            if cached is not None:
                result = dict(cached)
                result.update({
                    "cached": True,
                    "step_id": step_id,
                    "duration_ms": 0.0,
                    "completed_at": _utc_now().isoformat(),
                })
                return result

        # n8n-style "Retry On Fail": re-run the step up to max_tries on failure,
        # optionally waiting between attempts. Off by default (max_tries = 1).
        retry_on_fail = bool(config.get("retry_on_fail"))
        max_tries = max(1, min(int(config.get("max_tries", 3) or 3), 5)) if retry_on_fail else 1
        wait_secs = min(float(config.get("retry_wait_ms", 1000) or 0) / 1000.0, 30.0) if retry_on_fail else 0.0
        explicit_timeout = config.get("timeout")
        timeout = (
            float(explicit_timeout)
            if explicit_timeout is not None
            else None
            if step.get("type") in _DURABLE_ORCHESTRATION_NODE_TYPES
            else _STEP_TIMEOUT_SECS
        )

        result: dict = {"status": "failed", "error": "step did not run"}
        for attempt in range(max_tries):
            try:
                execution = self._execute_step(step, run, db)
                result = (
                    await execution
                    if timeout is None
                    else await asyncio.wait_for(execution, timeout=timeout)
                )
            except asyncio.TimeoutError:
                result = {"status": "failed", "error": f"Step {step_id} timed out"}
            except Exception as exc:
                logger.error("WorkflowRunner: step %s error: %s", step_id, exc, exc_info=True)
                result = {"status": "failed", "error": str(exc)}
            if result.get("status") == "completed":
                output_failure = _schema_validation_failure(
                    config.get("output_schema"),
                    result.get("output"),
                    code="output_schema_validation_failed",
                )
                if output_failure is not None:
                    result = output_failure
            if result.get("status") != "failed":
                break
            if attempt + 1 < max_tries and wait_secs > 0:
                await asyncio.sleep(wait_secs)
        if max_tries > 1:
            result["attempts"] = attempt + 1

        # Imported workflows are demos that reference live endpoints / credentials
        # often absent here (http calls, media generation, an LLM key, …). If an
        # *imported* node fails, skip it and let the run continue — the error is
        # preserved on the result, and any imported n8n/ComfyUI/Dify workflow runs
        # through to completion instead of dead-stopping on the first external
        # call. User-built nodes (no ``n8n`` marker) still fail normally; ``stop``
        # keeps its deliberate halt semantics.
        imported = bool((step.get("config") or {}).get("n8n") or (step.get("meta") or {}).get("source_tool"))
        if (
            result.get("status") == "failed"
            and imported
            and not config.get("strict_execution")
            and step.get("type") != "stop"
        ):
            result = {
                "status": "completed",
                "skipped": True,
                "output": "<skipped: needs a live endpoint / credentials in this environment>",
                "error": result.get("error"),
                "attempts": result.get("attempts"),
            }

        result["duration_ms"] = (time.monotonic() - start) * 1000
        result["completed_at"] = _utc_now().isoformat()
        result.setdefault("step_id", step_id)
        result.setdefault("inputs", input_snapshot)
        if fingerprint is not None:  # record so a future run can reuse this result
            result["fingerprint"] = fingerprint
        return result

    async def _execute_step(
        self, step: dict, run: WorkflowRun, db,
    ) -> dict:
        """Execute a single step, returning a result dict."""
        step_type = step.get("type", "tool")
        variables = dict(run.variables or {})
        # Explicit named inputs: bind config.inputs into the scope so the step's
        # own config can reference them by {{name}} (data-flow mapping).
        _bind_inputs(step.get("config") or {}, variables)
        entity_id = run.entity_id
        # user_id = whoever triggered this workflow run; MCP calls use
        # it to resolve personal OAuth tokens.
        user_id = run.started_by or ""
        runtime_context = runtime_workflow_run_context(run)

        if step_type == "stage":
            return await self._execute_stage_step(step, run, db)
        if step_type == "agent":
            if (step.get("config") or {}).get("batch"):
                return await self._execute_batch_llm_step(
                    step, variables, entity_id, user_id, runtime_context, db,
                )
            return await self._execute_agent_step(
                step, variables, entity_id, user_id, runtime_context, db,
            )
        elif step_type == "tool":
            return await self._execute_tool_step(
                step, variables, entity_id, user_id, runtime_context,
            )
        elif step_type == "condition":
            return await self._execute_condition_step(step, variables, run)
        elif step_type == "wait":
            return await self._execute_wait_step(step, run)
        elif step_type == "parallel":
            return await self._execute_parallel_step(step, variables, entity_id, run, db)
        elif step_type == "transform":
            return self._execute_transform_step(step, variables, run)
        elif step_type == "workflow_project":
            return await self._execute_workflow_project_step(step, variables, run, db)
        elif step_type == "workflow_action_grant":
            return await self._execute_workflow_action_grant_step(step, variables, run, db)
        elif step_type == "browser_effect":
            return await self._execute_browser_effect_step(step, variables)
        elif step_type == "notify":
            return await self._execute_notify_step(
                step, variables, entity_id, user_id, runtime_context,
            )
        elif step_type in ("llm", "classifier"):
            # llm / classifier are agent steps with no tools — reuse that path.
            if step_type == "llm" and (step.get("config") or {}).get("batch"):
                return await self._execute_batch_llm_step(
                    step, variables, entity_id, user_id, runtime_context, db,
                )
            return await self._execute_agent_step(
                step, variables, entity_id, user_id, runtime_context, db,
            )
        elif step_type == "connector":
            # A connector node is a single tool/MCP call — reuse the tool path.
            return await self._execute_tool_step(
                step, variables, entity_id, user_id, runtime_context,
            )
        elif step_type == "rag":
            return await self._execute_rag_step(
                step, variables, entity_id, user_id, runtime_context,
            )
        elif step_type == "http":
            return await self._execute_http_step(step, variables)
        elif step_type in ("media", "image", "video", "audio"):
            return await self._execute_media_step(
                step, variables, entity_id, user_id, runtime_context,
            )
        elif step_type == "code":
            return await self._execute_code_step(step, variables)
        elif step_type == "extract":
            if (step.get("config") or {}).get("batch"):
                return await self._execute_batch_extract_step(
                    step, variables, entity_id, user_id, runtime_context, db,
                )
            return await self._execute_extract_step(step, variables, entity_id, user_id, runtime_context, db)
        elif step_type == "filter":
            return self._execute_filter_step(step, variables)
        elif step_type == "aggregate":
            return self._execute_aggregate_step(step, variables)
        elif step_type == "datetime":
            return self._execute_datetime_step(step, variables)
        elif step_type == "split":
            return self._execute_split_step(step, variables)
        elif step_type == "limit":
            return self._execute_limit_step(step, variables)
        elif step_type == "respond":
            return self._execute_respond_step(step, variables)
        elif step_type == "sort":
            return self._execute_sort_step(step, variables)
        elif step_type == "dedupe":
            return self._execute_dedupe_step(step, variables)
        elif step_type == "stop":
            return self._execute_stop_step(step, variables)
        elif step_type == "extractfromfile":
            return self._execute_extractfromfile_step(step, variables)
        elif step_type == "subworkflow":
            return await self._execute_subworkflow_step(step, variables, run, db)
        elif step_type == "foreach_subworkflow":
            return await self._execute_foreach_subworkflow_step(step, variables, run, db)
        elif step_type == "loop":
            return await self._execute_loop_step(step, variables, entity_id, run, db)
        elif step_type == "switch":
            return self._execute_switch_step(step, variables, run)
        elif step_type == "merge":
            return self._execute_merge_step(step, variables, run)
        elif step_type == "end":
            # End is the workflow's business output. A named ``input`` binding
            # wins; otherwise pass through the most recently-recorded upstream
            # value. This keeps terminal nodes useful without requiring every
            # flow to repeat an explicit mapping.
            output = variables.get("input")
            if output is None:
                upstream = [value for key, value in variables.items() if not str(key).startswith("__")]
                output = upstream[-1] if upstream else None
            return {
                "status": "completed",
                "output": output if output is not None else (step.get("name") or step_type),
            }
        elif step_type in ("trigger", "webhook"):
            # Runtime trigger payload is the n8n-style first item. This matters
            # for chat/webhook flows where downstream nodes read fields such as
            # ``chatInput`` from the trigger output.
            output = dict(run.trigger_data or {}) or (step.get("name") or step_type)
            return {"status": "completed", "output": output}
        elif step_type == "note":
            # Canvas annotation — never part of the run; skip if ever reached.
            return {"status": "completed", "skipped": True, "output": ""}
        elif step_type == "unsupported":
            return self._execute_unsupported_step(step)
        else:
            return {"status": "failed", "error": f"Unknown step type: {step_type}"}

    async def _execute_stage_step(
        self,
        step: dict,
        run: WorkflowRun,
        db,
    ) -> dict:
        """Execute a checkpointed inline operation graph as one business stage."""
        stage_id = str(step["id"])
        config = step.get("config") if isinstance(step.get("config"), dict) else {}
        operations = [
            operation
            for operation in (config.get("operations") or [])
            if isinstance(operation, dict) and operation.get("id")
        ]
        operation_map = {str(operation["id"]): operation for operation in operations}
        entry_operation_id = str(config.get("entry_operation_id") or "")
        routes = config.get("routes") if isinstance(config.get("routes"), dict) else {}

        all_stage_state = deepcopy(
            (run.variables or {}).get("__stage_execution") or {}
        )
        state = deepcopy(all_stage_state.get(stage_id) or {})
        operation_results = {
            str(operation_id): _stage_operation_checkpoint(result)
            for operation_id, result in (state.get("operation_results") or {}).items()
            if isinstance(result, dict)
        }
        if state.get("status") == "failed":
            failed_operation_id = str(state.get("failed_operation_id") or "")
            operation_results.pop(failed_operation_id, None)

        synthetic_entry_id = f"__stage_entry__.{stage_id}"
        route_ids = {str(route_id) for route_id in routes}
        synthetic_steps = [
            {
                "id": synthetic_entry_id,
                "type": "trigger",
                "next": [entry_operation_id],
            },
            *operations,
            *[
                {"id": route_id, "type": "end", "next": []}
                for route_id in route_ids
            ],
        ]
        synthetic_workflow = SimpleNamespace(steps=synthetic_steps)
        variables = dict(run.variables or {})
        proxy = _StageRunProxy(
            run,
            variables=variables,
            operation_results=operation_results,
        )
        proxy.step_results[synthetic_entry_id] = {
            "status": "completed",
            "output": dict(run.trigger_data or {}),
        }

        def checkpoint(
            *,
            status: str,
            current_operation_id: str | None,
            pending_operation_ids: list[str],
            **extra: Any,
        ) -> None:
            nonlocal state, all_stage_state
            state = {
                **state,
                "status": status,
                "current_operation_id": current_operation_id,
                "pending_operation_ids": pending_operation_ids,
                "operation_results": deepcopy(operation_results),
                **extra,
            }
            all_stage_state = {
                **deepcopy(all_stage_state),
                stage_id: deepcopy(state),
            }
            parent_variables = dict(proxy.variables or {})
            parent_variables["__stage_execution"] = deepcopy(all_stage_state)
            run.variables = parent_variables
            run.trigger_data = deepcopy(proxy.trigger_data or {})
            run.current_step_id = stage_id

        def trace_node(operation: dict) -> dict:
            operation_config = dict(operation.get("config") or {})
            operation_config["chat_projection"] = "hidden"
            return {
                **operation,
                "id": f"{stage_id}.{operation['id']}",
                "name": operation.get("name") or operation["id"],
                "config": operation_config,
            }

        max_iterations = max(1, len(operations) * 3)
        for _ in range(max_iterations):
            runnable = self._find_runnable_steps(synthetic_workflow, proxy)
            runnable_operations = [
                candidate
                for candidate in runnable
                if str(candidate.get("id")) in operation_map
            ]
            if runnable_operations:
                pending_operation_ids = [
                    str(candidate["id"]) for candidate in runnable_operations
                ]
                operation = runnable_operations[0]
                operation_id = str(operation["id"])
                checkpoint(
                    status="running",
                    current_operation_id=operation_id,
                    pending_operation_ids=pending_operation_ids,
                    failed_operation_id=None,
                    paused_operation_id=None,
                )
                append_execution_trace(
                    run,
                    node=trace_node(operation),
                    status="running",
                )

                result = await self._execute_step_safe(operation, proxy, db)
                if result.get("status") == "failed" and _continues_on_error(operation):
                    result["continued"] = True
                self._record_step_result(operation, result, proxy)
                operation_results[operation_id] = _stage_operation_checkpoint(result)
                append_execution_trace(
                    run,
                    node=trace_node(operation),
                    status="skipped" if result.get("skipped") else result.get("status", "unknown"),
                    result=result,
                )

                if result.get("status") == "paused":
                    operation_config = (
                        operation.get("config")
                        if isinstance(operation.get("config"), dict)
                        else {}
                    )
                    checkpoint(
                        status="paused",
                        current_operation_id=operation_id,
                        pending_operation_ids=[operation_id],
                        paused_operation_id=operation_id,
                        failed_operation_id=None,
                    )
                    if db is not None:
                        await db.commit()
                    return {
                        **result,
                        "stage_id": stage_id,
                        "paused_operation_id": operation_id,
                        "wait_config": {
                            key: deepcopy(operation_config[key])
                            for key in (
                                "response_variable",
                                "options",
                                "approval_values",
                            )
                            if key in operation_config
                        },
                    }
                if result.get("status") == "failed" and not _continues_on_error(operation):
                    checkpoint(
                        status="failed",
                        current_operation_id=operation_id,
                        pending_operation_ids=[operation_id],
                        failed_operation_id=operation_id,
                        paused_operation_id=None,
                    )
                    if db is not None:
                        await db.commit()
                    return {
                        **result,
                        "stage_id": stage_id,
                        "failed_operation_id": operation_id,
                    }

                next_runnable = self._find_runnable_steps(synthetic_workflow, proxy)
                checkpoint(
                    status="running",
                    current_operation_id=None,
                    pending_operation_ids=[
                        str(candidate["id"])
                        for candidate in next_runnable
                        if str(candidate.get("id")) in operation_map
                    ],
                    failed_operation_id=None,
                    paused_operation_id=None,
                )
                if db is not None:
                    await db.commit()
                continue

            selected_routes = [
                str(candidate["id"])
                for candidate in runnable
                if str(candidate.get("id")) in route_ids
            ]
            if len(selected_routes) > 1:
                return {
                    "status": "failed",
                    "error": (
                        f"Stage {stage_id} selected multiple external routes: "
                        + ", ".join(selected_routes)
                    ),
                }

            selected_route = selected_routes[0] if selected_routes else None
            selected_target = routes.get(selected_route) if selected_route else None
            checkpoint(
                status="completed",
                current_operation_id=None,
                pending_operation_ids=[],
                selected_route=selected_route,
                selected_target=selected_target,
                failed_operation_id=None,
                paused_operation_id=None,
            )
            last_operation_id = next(
                (
                    str(operation["id"])
                    for operation in reversed(operations)
                    if str(operation["id"]) in operation_results
                ),
                None,
            )
            completed = {
                "status": "completed",
                "output": (
                    (proxy.variables or {}).get(last_operation_id)
                    if last_operation_id
                    else None
                ),
                "output_var": config.get("output_var"),
                "selected_route": selected_route,
            }
            if selected_route is not None:
                completed["next_override"] = (
                    [] if selected_target is None else [str(selected_target)]
                )
            return completed

        checkpoint(
            status="failed",
            current_operation_id=None,
            pending_operation_ids=[],
            failed_operation_id=None,
            paused_operation_id=None,
        )
        return {
            "status": "failed",
            "error": f"Stage {stage_id} exceeded maximum operation iterations",
        }

    # ── Agent step ───────────────────────────────────────────────────────

    async def _execute_batch_llm_step(
        self, step, variables, entity_id, user_id, runtime_context, db,
    ) -> dict:
        """Execute an imported item-stream LLM once per upstream item."""
        config = dict(step.get("config") or {})
        raw_items = config.get("items", variables.get("input"))
        items = _resolve_list(raw_items, variables)
        wrapper = str(config.pop("response_wrapper", "") or "")
        config.pop("batch", None)
        config.pop("items", None)
        outputs: list[Any] = []
        for item in items:
            item_scope = {**variables, "input": item, "item": item}
            if isinstance(item, dict):
                item_scope.update(item)
            result = await self._execute_agent_step(
                {**step, "config": config},
                item_scope,
                entity_id,
                user_id,
                runtime_context,
                db,
            )
            if result.get("status") != "completed":
                return result
            output: Any = result.get("output")
            if wrapper:
                wrapped: Any = output
                for part in reversed(wrapper.split(".")):
                    wrapped = {part: wrapped}
                output = wrapped
            outputs.append(output)
        return {
            "status": "completed",
            "output": outputs,
            "output_var": config.get("output_var"),
        }

    async def _execute_agent_step(
        self,
        step: dict,
        variables: dict,
        entity_id: str,
        user_id: str = "",
        runtime_context: dict[str, str | None] | None = None,
        db=None,
    ) -> dict:
        """Run agentic loop for an agent-type step.

        Config keys:
          - skill: skill ID or slug to invoke via Runtime skill boundary
          - prompt / input: text prompt (supports {{var}} templates)
          - system_prompt: override system prompt for the agentic loop
          - tools: list of tool names to make available
          - max_rounds: max agentic loop iterations (default 20)
          - temperature: LLM temperature (default 0.7)
          - model: LLM model override
        """
        config = step.get("config", {})
        runtime_context = runtime_context or {}
        resolved_agent_id = str(config.get("agent_id") or "").strip()
        if resolved_agent_id:
            resolved_agent_id = _render_template(resolved_agent_id, variables).strip()
            if "{{" in resolved_agent_id or "}}" in resolved_agent_id:
                resolved_agent_id = ""

        service_key = str(config.get("service_key") or "").strip()
        if service_key:
            service_key = _render_template(service_key, variables).strip()
            if "{{" in service_key or "}}" in service_key:
                service_key = ""

        if not resolved_agent_id and service_key and runtime_context.get("workspace_id"):
            if db is None:
                return {
                    "status": "failed",
                    "error": (
                        f"Workflow step {step.get('id') or step.get('name')!r} "
                        f"cannot resolve service_key {service_key!r} without a database session"
                    ),
                }
            from sqlalchemy import select as _select
            from packages.core.models.workspace import AgentSubscription as _AgentSubscription

            subscription = (
                await db.execute(
                    _select(_AgentSubscription).where(
                        _AgentSubscription.entity_id == entity_id,
                        _AgentSubscription.workspace_id == runtime_context.get("workspace_id"),
                        _AgentSubscription.service_key == service_key,
                        _AgentSubscription.status == "active",
                    ).limit(1)
                )
            ).scalar_one_or_none()
            if subscription is None:
                return {
                    "status": "failed",
                    "error": (
                        f"Workflow step {step.get('id') or step.get('name')!r} "
                        f"references missing active service_key {service_key!r}"
                    ),
                }
            resolved_agent_id = subscription.agent_id

        # If a skill is specified, delegate through the Runtime skill boundary.
        skill_ref = config.get("skill")
        output_schema = _resolve_structure(config.get("output_schema"), variables)
        if not isinstance(output_schema, dict):
            output_schema = None
        forced_tool_calls = _resolve_structure(
            config.get("forced_tool_calls"),
            variables,
        )
        if not isinstance(forced_tool_calls, list):
            forced_tool_calls = None
        elif not all(isinstance(call, dict) for call in forced_tool_calls):
            return {
                "status": "failed",
                "error": "Agent forced_tool_calls must be a list of objects",
            }
        if skill_ref:
            input_text = runtime_prompt_with_output_schema(
                _render_template(
                    config.get("input", config.get("prompt", "")),
                    variables,
                ),
                output_schema,
            )
            configured_tools = {str(name) for name in (config.get("tools") or []) if name}
            runtime_envelope = None
            allowed_tool_names = None
            runtime_request = runtime_request_for_surface_turn(
                surface=ChatSurface.WORKFLOW_AGENT_STEP,
                entity_id=entity_id,
                user_id=user_id or None,
                agent_id=resolved_agent_id or None,
                workspace_id=runtime_context.get("workspace_id"),
                conversation_id=runtime_context.get("conversation_id"),
                task_id=runtime_context.get("task_id"),
                message=input_text,
                legacy_path="ai.workflow_runner.skill_step",
            )
            if configured_tools:
                runtime_surface_result = runtime_prepare_named_tool_surface_for_turn(
                    runtime_request,
                    tool_names=configured_tools,
                )
                runtime_envelope = runtime_surface_result.envelope
                allowed_tool_names = runtime_surface_result.allowed_tool_names
            else:
                runtime_envelope = runtime_prepare_trace_envelope_for_turn(runtime_request)
            async with async_session() as db:
                skill_result = await runtime_invoke_skill(
                    db,
                    skill_ref,
                    entity_id,
                    input_text,
                    user_id=user_id or None,
                    agent_id=resolved_agent_id or None,
                    workspace_id=runtime_context.get("workspace_id"),
                    conversation_id=runtime_context.get("conversation_id"),
                    task_id=runtime_context.get("task_id"),
                    allowed_tool_names=allowed_tool_names,
                    runtime_envelope=runtime_envelope,
                    model=config.get("model") or None,
                    runtime_tool_context=runtime_workflow_tool_context_args(runtime_context),
                    output_schema=output_schema,
                    forced_tool_calls=forced_tool_calls,
                )
            if skill_result.get("error"):
                return await runtime_attach_and_persist_workflow_runner_result(
                    {"status": "failed", "error": skill_result["error"]},
                    runtime_envelope,
                )
            stop_reason = str(skill_result.get("stop_reason") or "").strip()
            configured_stop_reasons = config.get("fail_on_stop_reasons") or []
            if isinstance(configured_stop_reasons, str):
                configured_stop_reasons = [configured_stop_reasons]
            fail_on_stop_reasons = {
                str(reason).strip()
                for reason in configured_stop_reasons
                if str(reason or "").strip()
            }
            if stop_reason and stop_reason in fail_on_stop_reasons:
                failure = {
                    "status": "failed",
                    "error": str(skill_result.get("content") or stop_reason),
                    "stop_reason": stop_reason,
                    "usage": skill_result.get("usage"),
                    "tools_used": skill_result.get("tools_used", []),
                }
                return await runtime_attach_and_persist_workflow_runner_result(
                    failure,
                    runtime_envelope,
                )
            output = _parse_agent_output(
                skill_result.get("content", ""),
                config.get("output_format"),
            )
            # Store output in variables if output_var specified
            output_var = config.get("output_var")
            completed = {
                "status": "completed",
                "output": output,
                "output_var": output_var,
                "usage": skill_result.get("usage"),
                "tools_used": skill_result.get("tools_used", []),
            }
            return await runtime_attach_and_persist_workflow_runner_result(
                completed,
                runtime_envelope,
            )

        # No skill — run the prompt + tools through the Runtime Harness loop adapter.
        # An ``agent`` node can reference an existing manor Agent by id: its system
        # prompt, model and tool bindings (resolved via agent_id) come from that
        # Agent, so the node is just "pick an agent" rather than re-configuring one.
        agent_id = resolved_agent_id
        agent_row = None
        if agent_id and db is not None:
            from sqlalchemy import select as _select
            from packages.core.models.workspace import Agent as _Agent
            agent_row = (
                await db.execute(_select(_Agent).where(_Agent.id == agent_id))
            ).scalar_one_or_none()
        from packages.core.services.agent_runtime_config import (
            agent_runtime_config,
            agent_runtime_config_for,
        )
        agent_config = agent_runtime_config_for(agent_row)
        node_config = agent_runtime_config(config)

        default_system = "You are a helpful assistant completing a workflow step."
        if agent_row and agent_row.system_prompt:
            default_system = agent_row.system_prompt
        system_prompt = _render_template(config.get("system_prompt", default_system), variables)
        user_message = runtime_prompt_with_output_schema(
            _render_template(
                config.get("prompt", config.get("input", step.get("name", "Execute this step."))),
                variables,
            ),
            output_schema,
        )
        max_rounds = int(config.get("max_rounds", _AGENT_MAX_ROUNDS))
        temperature = (
            agent_config.temperature
            if agent_config.temperature is not None
            else (
                node_config.temperature
                if node_config.temperature is not None
                else _AGENT_TEMPERATURE
            )
        )

        # Resolve tools through Runtime prompt assembly so context blocks,
        # skill descriptors, tool filtering, and the trace envelope come from
        # one source of truth.
        tool_names = config.get("tools", [])
        runtime_request = runtime_request_for_surface_turn(
            surface=ChatSurface.WORKFLOW_AGENT_STEP,
            entity_id=entity_id,
            user_id=user_id or None,
            agent_id=agent_id or None,
            workspace_id=runtime_context.get("workspace_id"),
            conversation_id=runtime_context.get("conversation_id"),
            task_id=runtime_context.get("task_id"),
            message=user_message,
            legacy_path="ai.workflow_runner.agent_step",
        )
        try:
            appendix = await runtime_prepare_prompt_appendix_for_turn(
                db,
                request=runtime_request,
                active_user_message=user_message,
                configured_tool_names=tool_names,
            )
            tool_schemas = appendix.tool_schemas
            allowed_tool_names = appendix.allowed_tool_names
            runtime_envelope = appendix.envelope
            system_prompt = runtime_merge_prompt_appendix(system_prompt, appendix)
        except Exception:
            logger.debug("Workflow runtime prompt appendix failed; using tool surface fallback", exc_info=True)
            runtime_surface_result = runtime_prepare_named_tool_surface_for_turn(
                runtime_request,
                tool_names=tool_names,
            )
            tool_schemas = runtime_surface_result.tool_schemas
            allowed_tool_names = runtime_surface_result.allowed_tool_names
            runtime_envelope = runtime_surface_result.envelope

        result = await runtime_execute_workflow_agent_loop(
            runtime_envelope=runtime_envelope,
            system_prompt=system_prompt,
            user_message=user_message,
            tools=tool_schemas,
            entity_id=entity_id,
            agent_id=agent_id,
            user_id=user_id or None,
            workspace_id=runtime_context.get("workspace_id"),
            conversation_id=runtime_context.get("conversation_id"),
            task_id=runtime_context.get("task_id"),
            active_user_message=user_message,
            allowed_tool_names=allowed_tool_names,
            max_rounds=max_rounds,
            temperature=temperature,
            max_tokens=agent_config.max_tokens or node_config.max_tokens,
            model=agent_config.model or node_config.model,
            runtime_tool_context=runtime_workflow_tool_context_args(runtime_context),
            output_schema=output_schema,
            forced_tool_calls=forced_tool_calls,
        )

        output_var = config.get("output_var")
        output = _parse_agent_output(result.content, config.get("output_format"))
        completed = {
            "status": "completed",
            "output": output,
            "output_var": output_var,
            "usage": result.usage,
            "tools_used": result.tool_calls_made,
            "rounds": result.rounds,
        }
        return await runtime_attach_and_persist_workflow_runner_result(
            completed,
            runtime_envelope,
        )

    async def _execute_batch_extract_step(
        self, step, variables, entity_id, user_id, runtime_context, db,
    ) -> dict:
        """Run an imported Information Extractor once for every n8n item."""
        config = dict(step.get("config") or {})
        items = _resolve_list(config.pop("items", variables.get("input")), variables)
        config.pop("batch", None)
        outputs: list[Any] = []
        for item in items:
            item_scope = {**variables, "input": item, "item": item}
            if isinstance(item, dict):
                item_scope.update(item)
            result = await self._execute_extract_step(
                {**step, "config": config}, item_scope,
                entity_id, user_id, runtime_context, db,
            )
            if result.get("status") != "completed":
                return result
            outputs.append(result.get("output"))
        return {
            "status": "completed",
            "output": outputs,
            "output_var": config.get("output_var"),
        }

    # ── Tool step ────────────────────────────────────────────────────────

    async def _execute_tool_step(
        self,
        step: dict,
        variables: dict,
        entity_id: str,
        user_id: str = "",
        runtime_context: dict[str, str | None] | None = None,
    ) -> dict:
        """Execute a single tool for a tool-type step.

        Config keys:
          - tool: tool name to execute
          - args: dict of arguments (supports {{var}} templates in values)
          - output_var: variable name to store the result in
        """
        config = step.get("config", {})
        runtime_context = runtime_context or {}
        tool_name = config.get("tool", "")
        if not tool_name:
            # A freshly-imported connector (or any imported node) often has no
            # integration bound yet. Skip it so the rest of the run still
            # completes — the grey "skipped" state flags that it needs an
            # integration — rather than hard-failing the whole workflow. A bare
            # tool node the user added with no tool is still a real error.
            if step.get("type") == "connector" or config.get("n8n"):
                return {
                    "status": "completed",
                    "skipped": True,
                    "output": "<skipped: connect an integration to run this step>",
                }
            return {"status": "failed", "error": "No tool specified in step config"}

        raw_args = config.get("args", {})
        rendered_args = _resolve_structure(raw_args, variables)
        if not isinstance(rendered_args, dict):
            return {"status": "failed", "error": "Tool arguments must be an object"}
        runtime_request = runtime_request_for_surface_turn(
            surface=ChatSurface.WORKFLOW_AGENT_STEP,
            entity_id=entity_id,
            user_id=user_id or None,
            workspace_id=runtime_context.get("workspace_id"),
            conversation_id=runtime_context.get("conversation_id"),
            task_id=runtime_context.get("task_id"),
            message=str(config.get("prompt") or config.get("input") or step.get("name") or ""),
            legacy_path="ai.workflow_runner.tool_step",
        )
        tool_step_result = await runtime_execute_workflow_tool_step(
            request=runtime_request,
            tool_name=tool_name,
            arguments=rendered_args,
            active_user_message=str(config.get("prompt") or config.get("input") or step.get("name") or ""),
            workflow_context=runtime_context,
        )

        output_var = config.get("output_var")
        output: Any = tool_step_result.output
        if str(config.get("output_format") or "text").lower() == "json":
            if isinstance(output, str):
                try:
                    output = json.loads(output)
                except json.JSONDecodeError as exc:
                    return await runtime_attach_and_persist_workflow_runner_result(
                        {
                            "status": "failed",
                            "code": "tool_output_invalid_json",
                            "error": str(exc),
                        },
                        tool_step_result.envelope,
                    )
            if isinstance(output, dict) and (
                output.get("status") == "error" or output.get("error")
            ):
                return await runtime_attach_and_persist_workflow_runner_result(
                    {
                        "status": "failed",
                        "code": str(output.get("code") or "tool_step_failed"),
                        "error": str(output.get("error") or "Tool execution failed"),
                        "output": output,
                    },
                    tool_step_result.envelope,
                )
        completed = {
            "status": "completed",
            "output": output,
            "output_var": output_var,
        }
        return await runtime_attach_and_persist_workflow_runner_result(
            completed,
            tool_step_result.envelope,
        )

    # ── RAG step ─────────────────────────────────────────────────────────

    async def _execute_rag_step(
        self,
        step: dict,
        variables: dict,
        entity_id: str,
        user_id: str = "",
        runtime_context: dict[str, str | None] | None = None,
    ) -> dict:
        """Knowledge retrieval — a thin wrapper over the ``rag`` tool.

        Config keys:
          - query / question / input: the natural-language question
          - limit: max hits (optional)
          - workspace_id: scope to a specific workspace's knowledge (optional —
            node-level override; falls back to the run's workspace, else
            searches entity-wide)
          - output_var: variable to store the retrieved excerpts in
        """
        config = step.get("config", {})
        runtime_context = runtime_context or {}
        question = config.get("question") or config.get("query") or config.get("input") or ""
        args: dict = {"question": question}
        if config.get("limit") is not None:
            args["limit"] = config["limit"]
        # Node-level workspace pick wins; else inherit the run's workspace; else
        # entity-wide (standalone run with no workspace context).
        workspace_id = config.get("workspace_id") or runtime_context.get("workspace_id")
        if workspace_id:
            args["workspace_id"] = workspace_id
        # Restrict to specific knowledge collections (document groups) if picked.
        group_ids = config.get("group_ids") or config.get("net_ids")
        if group_ids:
            args["group_ids"] = group_ids
        tool_step = {
            "id": step.get("id"),
            "name": step.get("name"),
            "config": {"tool": "rag", "args": args, "output_var": config.get("output_var")},
        }
        return await self._execute_tool_step(
            tool_step, variables, entity_id, user_id, runtime_context,
        )

    # ── Loop step ────────────────────────────────────────────────────────

    async def _execute_loop_step(
        self, step: dict, variables: dict, entity_id: str, run: WorkflowRun, db,
    ) -> dict:
        """Iterate a collection, running inline sub-steps per item
        (n8n splitInBatches / ComfyUI loop).

        Config keys:
          - over / items: a variable name holding a list, or an inline list
          - item_var: variable name bound to the current item (default "item")
          - steps: inline sub-steps to run for each item
          - output_var: variable to collect per-iteration outputs
          - max_iterations: safety cap (default 100)
        """
        config = step.get("config", {})
        raw = config.get("over", config.get("items"))
        if isinstance(raw, str):
            items = variables.get(raw, [])
        elif isinstance(raw, list):
            items = raw
        else:
            items = [] if raw is None else [raw]
        if not isinstance(items, (list, tuple)):
            items = [items]

        item_var = config.get("item_var", "item")
        sub_steps = config.get("steps", [])
        cap = int(config.get("max_iterations", 100))

        outputs: list = []
        for i, item in enumerate(list(items)[:cap]):
            run.variables = {**(run.variables or {}), item_var: item, "index": i}
            for j, sub in enumerate(sub_steps):
                sub = dict(sub)
                sub.setdefault("id", f"{step['id']}_iter{i}_{j}")
                res = await self._execute_step_safe(sub, run, db)
                if res.get("status") == "failed":
                    return {
                        "status": "failed",
                        "error": f"Loop iteration {i} failed: {res.get('error')}",
                        "outputs": outputs,
                    }
                outputs.append(res.get("output"))

        return {"status": "completed", "output": outputs, "output_var": config.get("output_var")}

    # ── Media generation step (image / video / audio) ───────────────────

    async def _execute_media_step(
        self,
        step: dict,
        variables: dict,
        entity_id: str,
        user_id: str = "",
        runtime_context: dict[str, str | None] | None = None,
    ) -> dict:
        """Generate image / video / audio via the ``generate_file`` tool.

        The node ``type`` (image/video/audio) sets the kind; a generic ``media``
        node reads ``config.kind``. Config keys:
          - prompt / input: the generation prompt (supports {{var}})
          - kind: image | video | audio | document | presentation | ... (media node)
          - duration / resolution / aspect_ratio / first_frame_url / last_frame_url
            (video controls), name, output_var
        """
        config = step.get("config", {})
        step_type = step.get("type", "media")
        kind = step_type if step_type in ("image", "video", "audio") else config.get("kind", "image")

        args: dict = {
            "kind": kind,
            "prompt": _render_template(str(config.get("prompt") or config.get("input") or ""), variables),
        }
        # Optional per-node model override (picked from the catalog in the UI).
        # Blank falls through to the account's default model for this kind.
        if config.get("model"):
            args["model"] = str(config["model"]).strip()
        # Image controls (size / quality / reference) + video controls. The
        # generate_file tool's _merge_params hoists these into the per-kind
        # params. String values support {{var}} templates (e.g. a reference
        # image piped from an upstream step).
        for k in (
            "size", "quality", "reference_url", "reference_urls", "image_url",
            "input_fidelity", "save_to_knowledge",
            "duration", "resolution", "aspect_ratio", "first_frame_url",
            "last_frame_url", "name", "params",
        ):
            v = config.get(k)
            if v is not None:
                args[k] = _render_template(v, variables) if isinstance(v, str) else v

        tool_step = {
            "id": step.get("id"),
            "name": step.get("name"),
            "config": {"tool": "generate_file", "args": args, "output_var": config.get("output_var")},
        }
        return await self._execute_tool_step(
            tool_step, variables, entity_id, user_id, runtime_context,
        )

    # ── HTTP step ────────────────────────────────────────────────────────

    async def _execute_http_step(self, step: dict, variables: dict) -> dict:
        """Make an HTTP request (n8n httpRequest / Dify http-request).

        Config keys:
          - url (required): supports {{var}} templates
          - method: GET (default) | POST | PUT | PATCH | DELETE
          - headers: dict (values support templates)
          - body / json: request payload
          - timeout: seconds (default 30)
          - output_var: variable to store the parsed response in
        """
        import httpx

        config = step.get("config", {})
        if config.get("batch"):
            items = _resolve_list(config.get("items", variables.get("input")), variables)
            outputs: list[Any] = []
            single_config = dict(config)
            single_config.pop("batch", None)
            single_config.pop("items", None)
            for item in items:
                item_scope = {**variables, "input": item, "item": item}
                if isinstance(item, dict):
                    item_scope.update(item)
                result = await self._execute_http_step(
                    {**step, "config": single_config}, item_scope,
                )
                if result.get("status") != "completed":
                    return result
                outputs.append(result.get("output"))
            return {
                "status": "completed",
                "output": outputs,
                "output_var": config.get("output_var"),
            }

        url = _render_template(str(config.get("url", "")), variables)
        if not url:
            return {"status": "failed", "error": "No url specified in http step config"}

        method = str(config.get("method", "GET")).upper()
        headers = {
            k: _render_template(str(v), variables)
            for k, v in (config.get("headers") or {}).items()
        }
        query = {
            k: _render_template(str(v), variables)
            for k, v in (config.get("query") or {}).items()
        }
        timeout = float(config.get("timeout", 30))
        payload = config.get("json", config.get("body"))

        try:
            async with httpx.AsyncClient(timeout=timeout, follow_redirects=True) as http:
                resp = await http.request(
                    method,
                    url,
                    headers=headers or None,
                    params=query or None,
                    json=payload,
                )
        except httpx.HTTPError as exc:
            return {"status": "failed", "error": f"HTTP request failed: {exc}"}

        response_format = str(config.get("response_format") or "").lower()
        if response_format == "rss" and resp.is_success:
            try:
                articles = _parse_rss_feed(resp.text)
            except ElementTree.ParseError as exc:
                return {"status": "failed", "error": f"Invalid RSS/Atom response: {exc}"}
            if not articles:
                return {"status": "failed", "error": "RSS/Atom response contained no entries"}
            return {
                "status": "completed",
                "output": articles,
                "output_var": config.get("output_var"),
            }

        if response_format == "binary" and resp.is_success:
            return {
                "status": "completed",
                "output": {
                    "status_code": resp.status_code,
                    "content_type": resp.headers.get("content-type", "application/octet-stream"),
                    "body_base64": base64.b64encode(resp.content).decode("ascii"),
                },
                "output_var": config.get("output_var"),
            }

        try:
            data = resp.json()
        except (ValueError, UnicodeDecodeError):
            data = resp.text

        # n8n calls a text/file HTTP response ``data`` in several nodes while
        # Manor historically called it ``body``. Expose both names so imported
        # current-item expressions retain their meaning.
        output = {"status_code": resp.status_code, "body": data, "data": data}
        return {
            "status": "completed" if resp.is_success else "failed",
            "output": output,
            "output_var": config.get("output_var"),
            "error": None if resp.is_success else f"HTTP {resp.status_code}",
        }

    # ── Condition step ───────────────────────────────────────────────────

    async def _execute_condition_step(
        self, step: dict, variables: dict, run: WorkflowRun,
    ) -> dict:
        """Evaluate condition and determine branch.

        Config keys:
          - expression: e.g. ``score > 0.7``, ``status == "approved"``
        Step keys:
          - true_next: list of step IDs to follow if condition is true
          - false_next: list of step IDs to follow if condition is false
        """
        condition_scope = dict(variables)
        current_item = condition_scope.get("input")
        if isinstance(current_item, list) and len(current_item) == 1:
            current_item = current_item[0]
        if isinstance(current_item, dict):
            condition_scope.update(current_item)
        condition_met = self._evaluate_condition(step, condition_scope)

        if condition_met:
            next_steps = step.get("true_next", step.get("next", []))
        else:
            next_steps = step.get("false_next", [])

        return {
            "status": "completed",
            "output": current_item if (step.get("config") or {}).get("pass_input") else condition_met,
            "condition_result": condition_met,
            "next_override": next_steps,
        }

    # ── Durable orchestration state ──────────────────────────────────────

    @staticmethod
    def _workflow_project_output(project: Any) -> dict[str, Any]:
        return {
            "project_id": project.id,
            "project_type": project.project_type,
            "schema_version": project.schema_version,
            "current_stage": project.current_stage,
            "state": deepcopy(project.state or {}),
            "revision": project.revision,
            "last_run_id": project.last_run_id,
        }

    async def _execute_workflow_project_step(
        self,
        step: dict,
        variables: dict,
        run: WorkflowRun,
        db,
    ) -> dict:
        from packages.core.services.workflow_project_service import (
            WorkflowProjectConflict,
            create_workflow_project,
            get_workflow_project,
            patch_workflow_project,
        )

        config = step.get("config", {})
        operation = str(config.get("operation") or "get").strip().lower()
        output_var = config.get("output_var")
        state_schema = _resolve_structure(config.get("state_schema"), variables)
        if db is None:
            return {
                "status": "failed",
                "code": "workflow_project_failed",
                "error": "Workflow project node requires a database session",
            }
        if not run.workspace_id:
            return {
                "status": "failed",
                "code": "workflow_project_failed",
                "error": "Workflow project node requires a Workspace",
            }

        def validated_state(state: Any) -> dict[str, Any] | dict:
            if not isinstance(state, dict):
                return {
                    "status": "failed",
                    "code": "workflow_project_state_validation_failed",
                    "error": "Workflow project state must be an object",
                }
            failure = _schema_validation_failure(
                state_schema,
                state,
                code="workflow_project_state_validation_failed",
            )
            return failure or state

        try:
            project_type = str(
                _resolve_binding(config.get("project_type", ""), variables) or ""
            ).strip()
            project_id = str(
                _resolve_binding(config.get("project_id", ""), variables) or ""
            ).strip()

            if operation == "create":
                if not project_type:
                    raise ValueError("Workflow project create requires project_type")
                if not run.started_by:
                    raise ValueError("Workflow project create requires a triggering user")
                state = _resolve_structure(config.get("state", {}), variables)
                checked = validated_state(state)
                if checked.get("code") == "workflow_project_state_validation_failed":
                    return checked
                current_stage = str(
                    _resolve_binding(config.get("current_stage", "draft"), variables)
                    or "draft"
                ).strip()
                project = await create_workflow_project(
                    db,
                    entity_id=run.entity_id,
                    workspace_id=run.workspace_id,
                    project_type=project_type,
                    state=state,
                    created_by=run.started_by,
                    schema_version=max(1, int(config.get("schema_version") or 1)),
                    current_stage=current_stage,
                    last_run_id=run.id,
                )
            elif operation == "get":
                if not project_id:
                    raise ValueError("Workflow project get requires project_id")
                project = await get_workflow_project(
                    db,
                    project_id=project_id,
                    entity_id=run.entity_id,
                    workspace_id=run.workspace_id,
                )
                if project_type and project.project_type != project_type:
                    raise ValueError("Workflow project type does not match")
                checked = validated_state(deepcopy(project.state or {}))
                if checked.get("code") == "workflow_project_state_validation_failed":
                    return checked
            elif operation == "patch":
                if not project_id:
                    raise ValueError("Workflow project patch requires project_id")
                project = await get_workflow_project(
                    db,
                    project_id=project_id,
                    entity_id=run.entity_id,
                    workspace_id=run.workspace_id,
                )
                if project_type and project.project_type != project_type:
                    raise ValueError("Workflow project type does not match")
                expected_revision = int(
                    _resolve_binding(config.get("expected_revision"), variables)
                )
                patch = _resolve_structure(config.get("patch", {}), variables)
                if not isinstance(patch, dict):
                    raise ValueError("Workflow project patch must be an object")
                state = _merge_project_state(project.state or {}, patch)

                for update in config.get("list_upserts") or []:
                    if not isinstance(update, dict):
                        raise ValueError("Workflow project list_upserts entries must be objects")
                    path = str(_resolve_binding(update.get("path", ""), variables) or "")
                    key = str(update.get("key") or "id").strip() or "id"
                    item = _resolve_structure(update.get("item"), variables)
                    _upsert_project_list_item(_project_state_list(state, path), item, key)

                for update in config.get("list_reconciles") or []:
                    if not isinstance(update, dict):
                        raise ValueError("Workflow project list_reconciles entries must be objects")
                    path = str(_resolve_binding(update.get("path", ""), variables) or "")
                    key = str(update.get("key") or "id").strip() or "id"
                    ordered_keys = _resolve_structure(update.get("keys", []), variables)
                    replacements = _resolve_structure(update.get("items", []), variables)
                    if not isinstance(ordered_keys, list):
                        raise ValueError("Workflow project list_reconciles keys must be an array")
                    if not isinstance(replacements, list):
                        raise ValueError("Workflow project list_reconciles items must be an array")
                    _reconcile_project_list_items(
                        _project_state_list(state, path),
                        replacements,
                        key,
                        ordered_keys,
                    )

                for update in config.get("list_removes") or []:
                    if not isinstance(update, dict):
                        raise ValueError("Workflow project list_removes entries must be objects")
                    path = str(_resolve_binding(update.get("path", ""), variables) or "")
                    field = str(update.get("field") or "").strip()
                    values = _resolve_structure(update.get("values", []), variables)
                    if not isinstance(values, list):
                        raise ValueError("Workflow project list_removes values must be an array")
                    _remove_project_list_items(
                        _project_state_list(state, path),
                        field,
                        values,
                    )

                for update in config.get("list_appends") or []:
                    if not isinstance(update, dict):
                        raise ValueError("Workflow project list_appends entries must be objects")
                    path = str(_resolve_binding(update.get("path", ""), variables) or "")
                    key = str(update.get("key") or "id").strip() or "id"
                    resolved_items = _resolve_structure(update.get("items", []), variables)
                    items = resolved_items if isinstance(resolved_items, list) else [resolved_items]
                    target = _project_state_list(state, path)
                    for item in items:
                        _upsert_project_list_item(target, item, key)

                history_event = config.get("history_event")
                if history_event is not None:
                    event = _resolve_structure(history_event, variables)
                    if not isinstance(event, dict):
                        raise ValueError("Workflow project history_event must be an object")
                    event.setdefault("run_id", run.id)
                    event.setdefault(
                        "attempt_number",
                        run.effective_attempt_number,
                    )
                    event.setdefault("step_id", str(step.get("id") or ""))
                    event.setdefault("timestamp", _utc_now().isoformat())
                    _project_state_list(state, "history").append(event)

                checked = validated_state(state)
                if checked.get("code") == "workflow_project_state_validation_failed":
                    return checked
                current_stage_value = config.get("current_stage")
                current_stage = (
                    str(_resolve_binding(current_stage_value, variables) or "").strip()
                    if current_stage_value is not None
                    else None
                )
                allowed_stages = {
                    str(value) for value in (config.get("allowed_stages") or [])
                }
                if current_stage is not None and allowed_stages and current_stage not in allowed_stages:
                    raise ValueError(f"Workflow project stage is not allowed: {current_stage}")
                project = await patch_workflow_project(
                    db,
                    project_id=project_id,
                    entity_id=run.entity_id,
                    workspace_id=run.workspace_id,
                    expected_revision=expected_revision,
                    state=state,
                    current_stage=current_stage,
                    last_run_id=run.id,
                )
            else:
                raise ValueError(f"Unsupported Workflow project operation: {operation}")
            state = project.state if isinstance(project.state, dict) else {}
            _set_trusted_workflow_context(
                run,
                workflow_project_id=project.id,
                workflow_action_grant_id=state.get("capture_grant_id"),
                approved_plan_version=state.get("approved_plan_version"),
            )
            return {
                "status": "completed",
                "output": self._workflow_project_output(project),
                "output_var": output_var,
            }
        except WorkflowProjectConflict as exc:
            return {
                "status": "failed",
                "code": "workflow_project_conflict",
                "error": str(exc),
            }
        except Exception as exc:  # noqa: BLE001
            return {
                "status": "failed",
                "code": "workflow_project_failed",
                "error": str(exc),
            }

    async def _execute_workflow_action_grant_step(
        self,
        step: dict,
        variables: dict,
        run: WorkflowRun,
        db,
    ) -> dict:
        from packages.core.services.workflow_action_grant_service import (
            create_workflow_action_grant,
            revoke_workflow_action_grant,
        )
        from packages.core.services.workflow_project_service import get_workflow_project

        config = step.get("config", {})
        operation = str(config.get("operation") or "create").strip().lower()
        output_var = config.get("output_var")
        try:
            if db is None or not run.workspace_id or not run.started_by:
                raise ValueError("Workflow action grant requires database, Workspace, and user context")
            if operation == "revoke":
                grant_id = str(
                    _resolve_binding(config.get("grant_id", ""), variables) or ""
                ).strip()
                if not grant_id:
                    raise ValueError("Workflow action grant revoke requires grant_id")
                grant = await revoke_workflow_action_grant(
                    db,
                    grant_id=grant_id,
                    entity_id=run.entity_id,
                    workspace_id=run.workspace_id,
                )
            elif operation == "create":
                approval_step_id = str(config.get("approval_step_id") or "").strip()
                approval = (run.step_results or {}).get(approval_step_id) or {}
                if not (
                    approval_step_id
                    and approval.get("status") == "completed"
                    and approval.get("resumed") is True
                    and approval.get("wait_type") == "approval"
                    and approval.get("approved") is True
                    and approval.get("approved_by") == run.started_by
                ):
                    raise ValueError(
                        "Workflow action grant requires an explicit approved decision"
                    )
                project_id = str(
                    _resolve_binding(config.get("project_id", ""), variables) or ""
                ).strip()
                if not project_id:
                    raise ValueError("Workflow action grant create requires project_id")
                project = await get_workflow_project(
                    db,
                    project_id=project_id,
                    entity_id=run.entity_id,
                    workspace_id=run.workspace_id,
                )
                grant_type = str(
                    _resolve_binding(config.get("grant_type", ""), variables) or ""
                ).strip()
                if not grant_type:
                    raise ValueError("Workflow action grant create requires grant_type")
                scope = _resolve_structure(config.get("scope", {}), variables)
                failure = _schema_validation_failure(
                    config.get("scope_schema"),
                    scope,
                    code="workflow_action_grant_scope_validation_failed",
                )
                if failure is not None:
                    return failure
                if grant_type == "browser_capture":
                    state = project.state if isinstance(project.state, dict) else {}
                    approved_version = state.get("approved_plan_version")
                    if scope.get("approved_plan_version") != approved_version:
                        raise ValueError(
                            "Browser capture grant does not match the approved project plan"
                        )
                    plan = state.get("plan") if isinstance(state.get("plan"), dict) else {}
                    planned_scene_ids = {
                        str(value)
                        for value in (plan.get("scene_ids") or [])
                        if str(value or "").strip()
                    }
                    requested_scene_ids = {
                        str(value)
                        for value in (scope.get("scene_ids") or [])
                        if str(value or "").strip()
                    }
                    if not requested_scene_ids or not requested_scene_ids <= planned_scene_ids:
                        raise ValueError(
                            "Browser capture grant scenes are outside the approved project plan"
                        )
                grant = await create_workflow_action_grant(
                    db,
                    entity_id=run.entity_id,
                    workspace_id=run.workspace_id,
                    workflow_run_id=run.id,
                    project_id=project_id,
                    grant_type=grant_type,
                    scope=scope,
                    granted_by=str(approval["approved_by"]),
                    ttl_seconds=int(config.get("ttl_seconds") or 86400),
                )
            else:
                raise ValueError(f"Unsupported Workflow action grant operation: {operation}")
            if operation == "create":
                _set_trusted_workflow_context(
                    run,
                    workflow_project_id=grant.project_id,
                    workflow_action_grant_id=grant.id,
                    workflow_batch_capture=bool(
                        (grant.scope or {}).get("scene_ids")
                    ),
                    approved_plan_version=(grant.scope or {}).get(
                        "approved_plan_version"
                    ),
                )
            else:
                _set_trusted_workflow_context(
                    run,
                    workflow_action_grant_id=None,
                    workflow_batch_capture=None,
                )
            return {
                "status": "completed",
                "output": {
                    "grant_id": grant.id,
                    "project_id": grant.project_id,
                    "grant_type": grant.grant_type,
                    "scope": deepcopy(grant.scope or {}),
                    "granted_by": grant.granted_by,
                    "expires_at": grant.expires_at.isoformat(),
                    "revoked_at": grant.revoked_at.isoformat() if grant.revoked_at else None,
                },
                "output_var": output_var,
            }
        except Exception as exc:  # noqa: BLE001
            return {
                "status": "failed",
                "code": "workflow_action_grant_failed",
                "error": str(exc),
            }

    async def _execute_browser_effect_step(self, step: dict, variables: dict) -> dict:
        from packages.core.ai.runtime.browser_effects import (
            browser_effect_execution_decision,
            transition_browser_effect,
            validate_browser_effect_record,
        )

        config = step.get("config", {})
        operation = str(config.get("operation") or "decide").strip().lower()
        try:
            record = validate_browser_effect_record(
                _resolve_structure(config.get("record"), variables)
            )
            if operation == "transition":
                record = transition_browser_effect(
                    record,
                    str(_resolve_binding(config.get("target_status", ""), variables)),
                    evidence=_resolve_structure(config.get("evidence"), variables)
                    if config.get("evidence") is not None
                    else None,
                )
            elif operation != "decide":
                raise ValueError(f"Unsupported browser effect operation: {operation}")
            record_payload = {
                "effect_id": record.effect_id,
                "scene_id": record.scene_id,
                "action": record.action,
                "precondition": deepcopy(record.precondition),
                "expected_postcondition": deepcopy(record.expected_postcondition),
                "status": record.status.value,
                "evidence": [deepcopy(item) for item in record.evidence],
                "attempt_count": record.attempt_count,
            }
            return {
                "status": "completed",
                "output": {
                    "record": record_payload,
                    "decision": browser_effect_execution_decision(record).value,
                },
                "output_var": config.get("output_var"),
            }
        except Exception as exc:  # noqa: BLE001
            return {
                "status": "failed",
                "code": "browser_effect_failed",
                "error": str(exc),
            }

    # ── Wait step ────────────────────────────────────────────────────────

    async def _execute_wait_step(self, step: dict, run: WorkflowRun) -> dict:
        """Pause or delay execution — HITL approval, timer, or external event.

        Config keys:
          - wait_type: "approval" | "timer" | "event" (default "approval")
          - message: human-readable description of what we're waiting for
          - duration_seconds: for ``timer`` — how long to wait

        ``timer`` waits up to ``_INLINE_WAIT_CAP_SECS`` are slept inline so the
        flow continues automatically. Longer timers pause the run and enqueue a
        durable worker task that resumes it after the requested delay. Approval
        and event waits remain paused until an explicit resume signal arrives.
        """
        config = step.get("config", {})
        variables = dict(run.variables or {})
        wait_type = config.get("wait_type", "approval")
        message = _resolve_binding(
            config.get("message", f"Waiting for {wait_type}"),
            variables,
        )

        if wait_type == "timer":
            raw = _resolve_binding(
                config.get("duration_seconds", config.get("seconds", 0)),
                dict(run.variables or {}),
            )
            try:
                duration = float(raw)
            except (TypeError, ValueError):
                duration = 0.0
            if duration <= 0:
                return {
                    "status": "completed",
                    "output": "Waited 0s",
                    "wait_type": wait_type,
                }
            if 0 < duration <= _INLINE_WAIT_CAP_SECS:
                await asyncio.sleep(duration)
                return {
                    "status": "completed",
                    "output": f"Waited {duration:g}s",
                    "wait_type": wait_type,
                }
            # Too long to hold the run open inline. Pause now and let a worker
            # resume this exact run after the delay instead of occupying an API
            # request or worker process for minutes/hours.
            resume_at = _utc_now() + timedelta(seconds=duration)
            scheduled = self.enqueue_resume(run.id, duration)
            run.status = "paused"
            return {
                "status": "paused",
                "output": message,
                "wait_type": wait_type,
                "duration_seconds": duration,
                "resume_at": resume_at.isoformat(),
                "auto_resume_scheduled": scheduled,
            }

        run.status = "paused"
        result = {
            "status": "paused",
            "output": message,
            "wait_type": wait_type,
        }
        if config.get("review") is not None:
            result["review"] = _resolve_structure(config["review"], variables)
        if config.get("review_title") is not None:
            result["review_title"] = _resolve_binding(config["review_title"], variables)
        return result

    # ── Notify step ──────────────────────────────────────────────────────

    async def _execute_notify_step(
        self, step: dict, variables: dict, entity_id: str,
        user_id: str, runtime_context: dict,
    ) -> dict:
        """Send a notification through the platform's notification dispatcher.

        Config keys:
          - channel: optional channel name (e.g. "slack", "email"); omit to use
            the recipient's routing preferences (always includes the in-app bell).
          - message: notification body (templated)
        """
        config = step.get("config", {})
        message = _render_template(str(config.get("message") or ""), variables)
        title = step.get("name") or "Workflow notification"
        channel = config.get("channel")
        if not user_id:
            # No triggering user to target the in-app bell — degrade gracefully.
            return {"status": "completed", "output": "Notify skipped (no recipient)"}

        try:
            from packages.core.services.notify import notify as _notify
            await _notify(
                entity_id=entity_id,
                user_id=user_id,
                type="workflow",
                title=title,
                body=message or title,
                channels=[str(channel)] if channel else None,
                workspace_id=runtime_context.get("workspace_id"),
            )
        except Exception as exc:  # noqa: BLE001 — a failed notify must not fail the run
            logger.warning("notify step failed: %s", exc)
            return {"status": "completed", "output": f"Notify failed: {exc}"}
        return {"status": "completed", "output": f"Notified via {channel or 'default channels'}"}

    # ── Parallel step ────────────────────────────────────────────────────

    async def _execute_parallel_step(
        self, step: dict, variables: dict, entity_id: str,
        run: WorkflowRun, db,
    ) -> dict:
        """Run multiple sub-steps concurrently.

        Config keys:
          - steps: list of inline step dicts to execute in parallel
        """
        config = step.get("config", {})
        sub_steps = config.get("steps", [])
        if not sub_steps:
            return {"status": "completed", "output": "No sub-steps to execute"}

        tasks = []
        for sub in sub_steps:
            sub.setdefault("id", f"{step['id']}_sub_{len(tasks)}")
            tasks.append(self._execute_step_safe(sub, run, db))

        results = await asyncio.gather(*tasks)

        outputs = {}
        failed = []
        for sub, result in zip(sub_steps, results):
            sub_id = sub["id"]
            outputs[sub_id] = result.get("output")
            if result.get("status") == "failed":
                failed.append(sub_id)

        if failed:
            return {
                "status": "failed",
                "error": f"Parallel sub-steps failed: {', '.join(failed)}",
                "outputs": outputs,
            }

        return {
            "status": "completed",
            "output": outputs,
        }

    # ── Transform step ───────────────────────────────────────────────────

    def _execute_transform_step(
        self, step: dict, variables: dict, run: WorkflowRun,
    ) -> dict:
        """Update workflow variables via config.set mapping."""
        config = step.get("config", {})
        if "html_template" in config:
            source = _resolve_binding(config.get("items", "{{input}}"), variables)
            items = list(source) if isinstance(source, (list, tuple)) else [source]
            rendered = []
            for item in items:
                item_scope = {**variables, "input": item, "item": item}
                if isinstance(item, dict):
                    item_scope.update(item)
                rendered.append({"html": _render_template(str(config["html_template"]), item_scope)})
            output = rendered if isinstance(source, (list, tuple)) else rendered[0]
            return {"status": "completed", "output": output, "output_var": config.get("output_var")}
        if "markdown" in config:
            source = _resolve_binding(config.get("items", "{{input}}"), variables)
            item_scope = {**variables, "input": source, "item": source}
            if isinstance(source, dict):
                item_scope.update(source)
            rendered = _render_template(str(config.get("markdown") or ""), item_scope)
            if config.get("markdown_to_html"):
                from markdown import markdown
                rendered = markdown(rendered)
            return {
                "status": "completed",
                "output": {"data": rendered},
                "output_var": config.get("output_var"),
            }
        transforms = config.get("set", {})
        if "items" in config or "over" in config:
            raw = config.get("items", config.get("over"))
            source = _resolve_binding(raw, variables) if isinstance(raw, str) else raw
            items = list(source) if isinstance(source, (list, tuple)) else [source]
            mapped: list[dict[str, Any]] = []
            for item in items:
                item_scope = {**variables, "item": item}
                if isinstance(item, dict):
                    # n8n Set expressions refer to the current item's fields as
                    # $json.foo; imported mappings normalize those to {{foo}}.
                    item_scope.update(item)
                mapped_item = dict(item) if config.get("include_other_fields") and isinstance(item, dict) else {}
                mapped_item.update({
                    key: _resolve_structure(value, item_scope)
                    for key, value in transforms.items()
                })
                mapped.append(mapped_item)
            output: Any = mapped if isinstance(source, (list, tuple)) else (mapped[0] if mapped else {})
            return {
                "status": "completed",
                "output": output,
                "output_var": config.get("output_var"),
            }
        updated_vars = dict(run.variables or {})
        for key, value in transforms.items():
            updated_vars[key] = _resolve_structure(value, variables)
        run.variables = updated_vars
        return {"status": "completed", "output": updated_vars}

    # ── Switch step ──────────────────────────────────────────────────────

    def _execute_switch_step(
        self, step: dict, variables: dict, run: WorkflowRun,
    ) -> dict:
        """Multi-branch routing — first matching case wins, else default.

        Config keys:
          - cases: ``[{"expression": "status == \\"vip\\"", "next": ["stepA"]}, ...]``
          - default_next: list of step IDs when no case matches
        """
        config = step.get("config", {})
        for case in config.get("cases", []):
            expr = case.get("expression")
            if expr and self._evaluate_condition({"config": {"expression": expr}}, variables):
                return {
                    "status": "completed",
                    "output": expr,
                    "next_override": case.get("next", []),
                }
        default_next = config.get("default_next", step.get("next", []))
        return {"status": "completed", "output": "default", "next_override": default_next}

    # ── Merge step ───────────────────────────────────────────────────────

    def _execute_merge_step(
        self, step: dict, variables: dict, run: WorkflowRun,
    ) -> dict:
        """Aggregate several variables into one (Dify variable-aggregator).

        Config keys:
          - sources: list of variable names to combine
          - output_var: variable name to store the aggregate in
          - mode: "list" (default) | "dict"
        """
        config = step.get("config", {})
        sources = config.get("sources", [])
        mode = config.get("mode", "list")
        if mode == "combine_by_position":
            collections = [
                value if isinstance(value, list) else [value]
                for value in (variables.get(source) for source in sources)
            ]
            merged = []
            for index in range(max((len(value) for value in collections), default=0)):
                item: dict[str, Any] = {}
                for source, values in zip(sources, collections):
                    if index >= len(values):
                        continue
                    value = values[index]
                    if isinstance(value, dict):
                        item.update(value)
                    else:
                        item[source] = value
                merged.append(item)
        elif mode == "dict":
            merged: object = {s: variables.get(s) for s in sources}
        else:
            values = [variables.get(s) for s in sources]
            if config.get("flatten"):
                merged = []
                for value in values:
                    merged.extend(value if isinstance(value, list) else [value])
            else:
                merged = values
        return {
            "status": "completed",
            "output": merged,
            "output_var": config.get("output_var"),
        }

    # ── Code step ────────────────────────────────────────────────────────

    async def _execute_code_step(self, step: dict, variables: dict) -> dict:
        """Execute a Python, JavaScript, or Bash snippet in an ephemeral sandbox.

        The complete workflow variable bag is written to ``inputs.json``. Python
        and JavaScript snippets also receive it as a global ``inputs`` value; Bash
        receives ``WORKFLOW_INPUTS_FILE``. The sandbox is destroyed after every
        execution and has networking disabled unless the node explicitly opts in.
        """
        config = step.get("config", {})
        raw_language = str(config.get("language") or "python").strip().lower()
        language = {
            "py": "python",
            "python3": "python",
            "js": "javascript",
            "node": "javascript",
            "nodejs": "javascript",
            "sh": "bash",
            "shell": "bash",
        }.get(raw_language, raw_language)
        if language not in {"python", "javascript", "bash"}:
            return {
                "status": "failed",
                "error": f"Unsupported code language: {raw_language}",
            }

        code = str(config.get("code") or "")
        if not code.strip():
            return {"status": "failed", "error": "Code node has no code to execute"}

        sandbox_url = _workflow_sandbox_url()
        if not sandbox_url:
            return {
                "status": "failed",
                "error": (
                    "Code nodes require the Sandbox Service; configure "
                    "SANDBOX_SERVICE_URL before running this workflow"
                ),
            }

        requirements = config.get("requirements") or ""
        if isinstance(requirements, list):
            requirements = "\n".join(str(item) for item in requirements if item)
        requirements = str(requirements).strip()
        if requirements and config.get("allow_network") is not True:
            return {
                "status": "failed",
                "error": (
                    "Installing Python requirements needs Network access enabled "
                    "for this code node"
                ),
            }
        inputs_json = json.dumps(variables, ensure_ascii=False, default=str)
        files: dict[str, str] = {
            "SKILL.md": (
                "---\nname: workflow-code-node\n"
                "description: Ephemeral sandbox for one Manor workflow code node.\n---\n"
            ),
            "inputs.json": inputs_json,
        }
        if requirements:
            files["requirements.txt"] = requirements + "\n"

        if language == "python":
            files["user_code.py"] = code
            files["run.py"] = (
                "import json\n"
                "from pathlib import Path\n"
                "inputs = json.loads(Path('inputs.json').read_text(encoding='utf-8'))\n"
                "source = Path('user_code.py').read_text(encoding='utf-8')\n"
                "exec(compile(source, 'user_code.py', 'exec'), {'inputs': inputs, '__name__': '__main__'})\n"
            )
            command = "python run.py"
        elif language == "javascript":
            files["user_code.js"] = code
            files["run.js"] = (
                "const fs = require('fs');\n"
                "globalThis.inputs = JSON.parse(fs.readFileSync('inputs.json', 'utf8'));\n"
                "require('./user_code.js');\n"
            )
            command = "node run.js"
        else:
            files["run.sh"] = code
            command = "WORKFLOW_INPUTS_FILE=inputs.json bash run.sh"

        try:
            code_timeout = max(
                1,
                min(int(config.get("code_timeout") or config.get("timeout") or 60), 300),
            )
        except (TypeError, ValueError):
            code_timeout = 60

        sandbox_id: str | None = None
        try:
            from packages.core.services.sandbox_sdk import SandboxClient

            async with SandboxClient(
                base_url=sandbox_url,
                timeout=float(code_timeout) + 60.0,
            ) as client:
                created = await client.create_from_files(
                    skill_name=f"workflow-code-{re.sub(r'[^a-zA-Z0-9_-]', '-', str(step.get('id') or 'node'))}",
                    files=files,
                    env={},
                    allowed_sensitive_keys=[],
                    auto_install=True,
                    config={
                        "network": "bridge" if config.get("allow_network") is True else "none",
                        "memory": "512m",
                        "cpus": 1.0,
                        "pids_limit": 128,
                    },
                )
                sandbox_id = created.sandbox_id
                executed = await client.exec(
                    sandbox_id=sandbox_id,
                    command=command,
                    timeout=code_timeout,
                    workdir=created.workdir,
                )
                try:
                    await client.destroy(sandbox_id)
                    sandbox_id = None
                except Exception:  # noqa: BLE001 - cleanup is retried in finally
                    logger.warning("workflow code sandbox cleanup failed", exc_info=True)

            if executed.exit_code != 0:
                return {
                    "status": "failed",
                    "error": executed.stderr or f"Code exited with status {executed.exit_code}",
                    "output": executed.stdout,
                    "stderr": executed.stderr,
                    "exit_code": executed.exit_code,
                }

            stdout = executed.stdout or ""
            output_format = str(config.get("output_format") or "auto").lower()
            output: Any = stdout
            if output_format in {"auto", "json"} and stdout.strip():
                try:
                    output = json.loads(stdout)
                except json.JSONDecodeError:
                    if output_format == "json":
                        return {
                            "status": "failed",
                            "error": "Code output is not valid JSON",
                            "output": stdout,
                            "stderr": executed.stderr,
                        }
            return {
                "status": "completed",
                "output": output,
                "stderr": executed.stderr,
                "exit_code": executed.exit_code,
                "sandboxed": True,
                "language": language,
                "output_var": config.get("output_var"),
            }
        except Exception as exc:  # noqa: BLE001 - convert sandbox failures to node failures
            return {"status": "failed", "error": f"Sandbox execution failed: {exc}"}
        finally:
            if sandbox_id:
                try:
                    from packages.core.services.sandbox_sdk import SandboxClient

                    async with SandboxClient(base_url=sandbox_url, timeout=30.0) as client:
                        await client.destroy(sandbox_id)
                except Exception:  # noqa: BLE001 - best-effort leak prevention
                    logger.warning(
                        "workflow code sandbox %s could not be destroyed",
                        sandbox_id,
                        exc_info=True,
                    )

    # ── Information Extractor (n8n) — LLM → structured JSON ──────────────

    async def _execute_extract_step(
        self, step, variables, entity_id, user_id, runtime_context, db,
    ) -> dict:
        """Extract structured fields from text with a tool-less LLM, parsing the
        reply as JSON. config: {model?, input/text, schema/fields}."""
        from packages.core.services.skill_bundle import extract_json_object

        config = step.get("config", {})
        html_fields = config.get("html_extract")
        if isinstance(html_fields, list):
            source = variables.get("input")
            sources = list(source) if isinstance(source, (list, tuple)) else [source]
            try:
                outputs = [_html_extract_fields(item, html_fields) for item in sources]
            except Exception as exc:
                return {"status": "failed", "error": f"HTML extraction failed: {exc}"}
            return {
                "status": "completed",
                "output": outputs if isinstance(source, (list, tuple)) else outputs[0],
                "output_var": config.get("output_var"),
            }

        schema = config.get("schema") or config.get("fields") or ""
        if isinstance(schema, (dict, list)):
            schema = json.dumps(schema, ensure_ascii=False)
        source = config.get("input") or config.get("text") or config.get("prompt") or ""
        extractor_instruction = (
            "You are an information-extraction engine. Extract the requested fields "
            "from the user's text and respond with ONLY a JSON object — no prose, no "
            "code fences. Use null for any field that is absent.\nFields to extract:\n"
            + str(schema)
        )
        imported_instruction = str(config.get("system_prompt") or "").strip()
        system = (
            f"{imported_instruction}\n\n{extractor_instruction}"
            if imported_instruction else extractor_instruction
        )
        llm_step = {
            "id": step.get("id"), "type": "llm", "name": step.get("name"),
            "config": {"model": config.get("model"), "system_prompt": system, "prompt": source},
        }
        res = await self._execute_agent_step(
            llm_step, variables, entity_id, user_id, runtime_context, db,
        )
        if res.get("status") != "completed":
            return res
        parsed = extract_json_object(str(res.get("output") or ""))
        output: Any = parsed if parsed else res.get("output")
        wrapper = str(config.get("response_wrapper") or "")
        if wrapper:
            for part in reversed(wrapper.split(".")):
                output = {part: output}
        return {
            "status": "completed",
            "output": output,
            "output_var": config.get("output_var"),
        }

    # ── Filter (n8n) — keep list items matching a condition ──────────────

    def _execute_filter_step(self, step, variables) -> dict:
        """Keep items of a list for which a per-item condition holds.
        config: {items, item_var?, condition}. The condition uses the same
        evaluator as the IF node, with the item bound to ``item_var``."""
        config = step.get("config", {})
        items = _resolve_list(config.get("items", config.get("over")), variables)
        field = config.get("field")
        operation = str(config.get("operation") or "").lower()
        if field and operation in {"after", "before"}:
            threshold_raw = config.get("threshold")
            threshold = (
                _parse_dt(str(_resolve_binding(threshold_raw, variables)))
                if threshold_raw not in (None, "")
                else _utc_now() + timedelta(days=float(config.get("relative_days", 0) or 0))
            )
            if threshold is None:
                return {"status": "failed", "error": "Filter has an invalid date threshold"}
            if threshold.tzinfo is None:
                threshold = threshold.replace(tzinfo=timezone.utc)
            kept = []
            for item in items:
                value = _dig(item, field)
                parsed = _parse_dt(str(value)) if value not in (None, "") else None
                if parsed is None:
                    continue
                if parsed.tzinfo is None:
                    parsed = parsed.replace(tzinfo=timezone.utc)
                if (operation == "after" and parsed > threshold) or (
                    operation == "before" and parsed < threshold
                ):
                    kept.append(item)
            return {"status": "completed", "output": kept, "output_var": config.get("output_var")}
        item_var = config.get("item_var", "item")
        expr = str(config.get("condition") or config.get("expression") or "").strip()
        if not expr:
            return {"status": "completed", "output": items, "output_var": config.get("output_var")}
        kept = []
        for it in items:
            scope = {**variables, item_var: it}
            try:
                if self._eval_bool_expr(expr, scope):
                    kept.append(it)
            except Exception:  # noqa: BLE001 — a bad item never kills the filter
                pass
        return {"status": "completed", "output": kept, "output_var": config.get("output_var")}

    # ── Aggregate (n8n) — reduce a list to a single value ────────────────

    def _execute_aggregate_step(self, step, variables) -> dict:
        """Reduce a list. config: {items, operation, field?, separator?}.
        operation: count | sum | avg | min | max | join | first | last | collect."""
        config = step.get("config", {})
        items = _resolve_list(config.get("items", config.get("over")), variables)
        field = config.get("field")
        op = str(config.get("operation") or "collect").lower()
        fields = config.get("fields") or []
        max_field_chars = int(config.get("max_field_chars", 0) or 0)
        if fields:
            vals = []
            for item in items:
                selected = {name: _dig(item, name) for name in fields}
                if max_field_chars:
                    selected = {
                        key: value[:max_field_chars] if isinstance(value, str) else value
                        for key, value in selected.items()
                    }
                vals.append(selected)
        else:
            vals = [_dig(it, field) for it in items]
        nums = [v for v in vals if isinstance(v, (int, float)) and not isinstance(v, bool)]
        if op == "count":
            out: Any = len(items)
        elif op == "sum":
            out = sum(nums)
        elif op == "avg":
            out = (sum(nums) / len(nums)) if nums else None
        elif op == "min":
            out = min(nums) if nums else None
        elif op == "max":
            out = max(nums) if nums else None
        elif op == "join":
            out = str(config.get("separator", ", ")).join(str(v) for v in vals if v is not None)
        elif op == "first":
            out = vals[0] if vals else None
        elif op == "last":
            out = vals[-1] if vals else None
        else:  # collect
            out = vals
        if config.get("wrap_key"):
            out = {str(config["wrap_key"]): out}
        return {"status": "completed", "output": out, "output_var": config.get("output_var")}

    # ── Date & Time (n8n) — produce / format / shift a datetime ──────────

    def _execute_datetime_step(self, step, variables) -> dict:
        """config: {operation, value?, format?, amount?, unit?}.
        operation: now | format | add | subtract. ``value`` (templated) defaults
        to now; ``format`` is a strftime pattern (ISO if omitted)."""
        config = step.get("config", {})
        op = str(config.get("operation") or "now").lower()
        fmt = config.get("format")
        value = _render_template(str(config.get("value") or ""), variables)
        base = _parse_dt(value) or _utc_now()
        if op in ("add", "subtract"):
            try:
                amount = float(config.get("amount") or 0)
            except (TypeError, ValueError):
                amount = 0.0
            unit = str(config.get("unit") or "days").lower()
            key = {"second": "seconds", "seconds": "seconds", "minute": "minutes",
                   "minutes": "minutes", "hour": "hours", "hours": "hours",
                   "day": "days", "days": "days", "week": "weeks", "weeks": "weeks"}.get(unit, "days")
            delta = timedelta(**{key: amount})
            base = base + delta if op == "add" else base - delta
        try:
            out = base.strftime(fmt) if fmt else base.isoformat()
        except (ValueError, TypeError):
            out = base.isoformat()
        return {"status": "completed", "output": out, "output_var": config.get("output_var")}

    # ── Split Out (n8n) — explode a list / delimited string ──────────────

    def _execute_split_step(self, step, variables) -> dict:
        """Turn a list field (or a delimited string) into a flat list.
        config: {items, field?, separator?}. A list of objects + ``field`` plucks
        (and flattens) that field from each; a string + ``separator`` splits it."""
        config = step.get("config", {})
        src = config.get("items", config.get("field"))
        val = _resolve_binding(src, variables) if isinstance(src, str) else src
        sep = config.get("separator")
        field = config.get("field")
        if isinstance(val, dict) and field:
            val = _dig(val, field)
        if isinstance(val, str):
            out: list = [p.strip() for p in val.split(str(sep))] if sep is not None else [val]
        elif isinstance(val, (list, tuple)):
            if field and any(isinstance(item, dict) and field in item for item in val):
                out = []
                for it in val:
                    v = _dig(it, field)
                    out.extend(v) if isinstance(v, (list, tuple)) else out.append(v)
            else:
                out = list(val)
        else:
            out = [] if val is None else [val]
        if field and config.get("preserve_field"):
            out = [{field: item} for item in out]
        return {"status": "completed", "output": out, "output_var": config.get("output_var")}

    # ── Limit (n8n) — cap a list ─────────────────────────────────────────

    def _execute_limit_step(self, step, variables) -> dict:
        """Keep the first (or last) N items of a list.
        config: {items, max, keep?}  keep: 'first' (default) | 'last'."""
        config = step.get("config", {})
        items = _resolve_list(config.get("items", config.get("over")), variables)
        try:
            n = max(0, int(config.get("max", config.get("limit", 1)) or 0))
        except (TypeError, ValueError):
            n = len(items)
        keep = str(config.get("keep", "first")).lower()
        out = items[-n:] if keep == "last" else items[:n]
        return {"status": "completed", "output": out, "output_var": config.get("output_var")}

    # ── Respond to Webhook (n8n) — set the inbound HTTP response ──────────

    def _execute_respond_step(self, step, variables) -> dict:
        """Capture the HTTP response a webhook-triggered run should return.
        config: {body, status_code?}. Stored under the reserved ``__response``
        variable, which the webhook endpoint returns to the caller."""
        config = step.get("config", {})
        body_raw = _render_template(str(config.get("body") or config.get("respondWith") or ""), variables)
        body: Any = body_raw
        if body_raw.strip().startswith(("{", "[")):
            try:
                body = json.loads(body_raw)
            except (ValueError, TypeError):
                body = body_raw
        try:
            status_code = int(config.get("status_code", config.get("status", 200)) or 200)
        except (TypeError, ValueError):
            status_code = 200
        resp = {"status": status_code, "body": body}
        return {"status": "completed", "output": resp, "output_var": "__response"}

    # ── Sort / Remove Duplicates / Stop and Error / Extract From File ────

    def _execute_sort_step(self, step, variables) -> dict:
        """Sort a list. config: {items, field?, order?}  order: asc | desc."""
        config = step.get("config", {})
        items = _resolve_list(config.get("items", config.get("over")), variables)
        field = config.get("field")
        desc = str(config.get("order", "asc")).lower().startswith("desc")
        try:
            out = sorted(items, key=lambda it: ((_dig(it, field) if field else it) is None,
                                                _dig(it, field) if field else it), reverse=desc)
        except TypeError:
            out = items  # mixed/uncomparable types — leave as-is
        return {"status": "completed", "output": out, "output_var": config.get("output_var")}

    def _execute_dedupe_step(self, step, variables) -> dict:
        """Remove duplicate list items. config: {items, field?} — by ``field``
        (or the whole item) identity."""
        config = step.get("config", {})
        items = _resolve_list(config.get("items", config.get("over")), variables)
        field = config.get("field")
        seen: set = set()
        out: list = []
        for it in items:
            k = _dig(it, field) if field else it
            try:
                h = k if isinstance(k, (str, int, float, bool, type(None))) else json.dumps(k, sort_keys=True, default=str)
            except (TypeError, ValueError):
                h = str(k)
            if h not in seen:
                seen.add(h)
                out.append(it)
        return {"status": "completed", "output": out, "output_var": config.get("output_var")}

    def _execute_stop_step(self, step, variables) -> dict:
        """n8n Stop And Error — deliberately fail the run with a message."""
        msg = _render_template(str(step.get("config", {}).get("message") or "Stopped by workflow"), variables)
        return {"status": "failed", "error": msg}

    def _execute_extractfromfile_step(self, step, variables) -> dict:
        """Parse text content as JSON or CSV into structured data.
        config: {input/text, format?}  format: auto (default) | json | csv.
        (Binary/PDF extraction is out of scope; this handles textual content.)"""
        config = step.get("config", {})
        raw = config.get("input", config.get("text", ""))
        content = _resolve_binding(raw, variables) if isinstance(raw, str) else raw
        fmt = str(config.get("format", "auto")).lower()
        if fmt in {"xls", "xlsx"}:
            encoded = content.get("body_base64") if isinstance(content, dict) else content
            if not isinstance(encoded, str) or not encoded:
                return {"status": "failed", "error": "spreadsheet input has no binary content"}
            try:
                from io import BytesIO
                from openpyxl import load_workbook

                workbook = load_workbook(BytesIO(base64.b64decode(encoded)), read_only=True, data_only=True)
                sheet = workbook.active
                rows = list(sheet.iter_rows(values_only=True))
                if not rows:
                    output: list[dict[str, Any]] = []
                else:
                    headers = [str(value or "").strip() for value in rows[0]]
                    output = [
                        {header: value for header, value in zip(headers, row) if header}
                        for row in rows[1:]
                        if any(value not in (None, "") for value in row)
                    ]
                max_rows = int(config.get("max_rows") or 0)
                if max_rows > 0:
                    output = output[:max_rows]
                return {"status": "completed", "output": output, "output_var": config.get("output_var")}
            except Exception as exc:
                return {"status": "failed", "error": f"invalid spreadsheet: {exc}"}
        text = str(content or "")
        s = text.strip()
        if fmt in ("json", "auto") and s[:1] in ("{", "["):
            try:
                return {"status": "completed", "output": json.loads(s), "output_var": config.get("output_var")}
            except (ValueError, TypeError):
                if fmt == "json":
                    return {"status": "failed", "error": "invalid JSON"}
        if fmt in ("csv", "auto") and s:
            import csv
            import io
            try:
                rows = list(csv.DictReader(io.StringIO(s)))
                if rows:
                    return {"status": "completed", "output": rows, "output_var": config.get("output_var")}
            except (csv.Error, ValueError):
                pass
        return {"status": "completed", "output": text, "output_var": config.get("output_var")}

    # ── Sub-workflow step (n8n Execute Workflow) ─────────────────────────

    async def _execute_subworkflow_step(
        self, step: dict, variables: dict, run: WorkflowRun, db,
    ) -> dict:
        """Run another workflow inline and return its result.

        Config: ``workflow_id`` (the target; supports {{var}}). The target is
        seeded with this step's named Inputs (already resolved into ``variables``)
        and its final variable bag becomes this step's output, so downstream refs
        reach a sub-workflow value via ``{{this_step.var}}``. Guards prevent
        self-calls and runaway recursion.
        """
        from sqlalchemy import or_, select

        config = step.get("config", {})
        target_id = _render_template(str(config.get("workflow_id") or ""), variables).strip()
        if not target_id:
            return {"status": "failed", "error": "subworkflow step has no workflow_id"}
        if target_id == run.workflow_id:
            return {"status": "failed", "error": "subworkflow cannot call itself"}
        depth = int((run.trigger_data or {}).get("_subworkflow_depth", 0))
        if depth >= _MAX_SUBWORKFLOW_DEPTH:
            return {"status": "failed", "error": f"subworkflow depth limit ({_MAX_SUBWORKFLOW_DEPTH}) exceeded"}

        res = await db.execute(
            select(WorkflowDefinition).where(
                WorkflowDefinition.entity_id == run.entity_id,
                or_(
                    WorkflowDefinition.id == target_id,
                    WorkflowDefinition.name == target_id,
                ),
            ).limit(1)
        )
        target = res.scalar_one_or_none()
        if target is None:
            return {"status": "failed", "error": f"subworkflow {target_id} not found"}
        if target.id == run.workflow_id:
            return {"status": "failed", "error": "subworkflow cannot call itself"}

        # Seed the child with the target's defaults + this node's named inputs
        # (already resolved + type-coerced into `variables` by _bind_inputs).
        seed = dict(target.variables or {})
        for item in config.get("inputs") or []:
            if isinstance(item, dict):
                k = str(item.get("key") or item.get("name") or "").strip()
                if k and k in variables:
                    seed[k] = variables[k]

        from packages.core.services.workflow_service import (
            _snapshot_display_metadata,
            entry_step_id,
        )

        steps = target.steps or []
        child_entry = entry_step_id(steps)
        if child_entry is None:
            return {
                "status": "failed",
                "error": "subworkflow requires an explicit trigger or webhook entry node",
            }
        target_fingerprint = workflow_definition_fingerprint(target)
        target_snapshot = build_definition_snapshot(
            target,
            fingerprint=target_fingerprint,
        )
        sub_run_id = generate_ulid()
        sub_run = WorkflowRun(
            id=sub_run_id,
            workflow_id=target.id,
            entity_id=run.entity_id,
            workspace_id=run.workspace_id,
            lineage_root_run_id=sub_run_id,
            lineage_is_legacy=False,
            attempt_number=1,
            status="running",
            current_step_id=child_entry,
            variables=seed,
            step_results={},
            definition_snapshot=target_snapshot,
            execution_trace=[],
            trigger_data={
                **_snapshot_display_metadata(target_snapshot),
                "_subworkflow_depth": depth + 1,
                "parent_run_id": run.id,
                "parent_step_id": step.get("id"),
                "_workflow_definition_version": target_snapshot["version"],
                "_workflow_definition_fingerprint": target_fingerprint,
                "_workflow_runtime_context": _child_workflow_context(run, seed),
            },
            started_by=run.started_by,
            started_at=_utc_now(),
        )
        db.add(sub_run)
        await db.flush()
        sub_id = sub_run.id
        # Persist so the nested runner (its own session) can load the child run.
        await db.commit()

        await WorkflowRunner(cache_index=self._cache_index).run(sub_id)

        # READ COMMITTED: a fresh read sees the nested run's committed result.
        await db.refresh(sub_run)
        if sub_run.status == "failed":
            return {"status": "failed", "error": sub_run.error or "subworkflow failed", "subrun_id": sub_id}
        if sub_run.status == "paused":
            return {
                "status": "paused",
                "output": "Waiting for subworkflow to resume",
                "subrun_id": sub_id,
            }
        return {
            "status": "completed",
            "output": dict(sub_run.variables or {}),
            "output_var": config.get("output_var"),
            "subrun_id": sub_id,
        }

    @staticmethod
    def _foreach_public_items(items: list[dict]) -> list[dict]:
        return [
            {
                key: item.get(key)
                for key in (
                    "key",
                    "index",
                    "status",
                    "subrun_id",
                    "attempts",
                    "output",
                    "error",
                )
            }
            for item in items
        ]

    @staticmethod
    def _set_foreach_progress(run: WorkflowRun, progress_key: str, progress: dict) -> None:
        updated = dict(run.variables or {})
        updated[progress_key] = deepcopy(progress)
        run.variables = updated

    async def _execute_foreach_subworkflow_step(
        self,
        step: dict,
        variables: dict,
        run: WorkflowRun,
        db,
    ) -> dict:
        """Durably map a child WorkflowRun over stable, ordered input items."""
        from sqlalchemy import or_, select

        config = step.get("config", {})
        target_id = _render_template(str(config.get("workflow_id") or ""), variables).strip()
        if not target_id:
            return {"status": "failed", "error": "foreach_subworkflow has no workflow_id"}
        if target_id == run.workflow_id:
            return {"status": "failed", "error": "foreach_subworkflow cannot call itself"}
        depth = int((run.trigger_data or {}).get("_subworkflow_depth", 0))
        if depth >= _MAX_SUBWORKFLOW_DEPTH:
            return {
                "status": "failed",
                "error": f"subworkflow depth limit ({_MAX_SUBWORKFLOW_DEPTH}) exceeded",
            }

        raw_items = config.get("over", config.get("items"))
        if isinstance(raw_items, str) and raw_items in variables:
            resolved_items = variables[raw_items]
        else:
            resolved_items = _resolve_binding(raw_items, variables)
        if not isinstance(resolved_items, list):
            return {"status": "failed", "error": "foreach_subworkflow items must be a list"}

        item_var = str(config.get("item_var") or "item").strip() or "item"
        item_key_field = str(config.get("item_key") or "id").strip() or "id"
        concurrency = max(1, min(int(config.get("concurrency") or 1), 8))
        progress_key = f"__foreach_subworkflow__{step['id']}"
        input_fingerprint = hashlib.sha256(
            json.dumps(resolved_items, sort_keys=True, default=str).encode("utf-8")
        ).hexdigest()

        existing = (run.variables or {}).get(progress_key)
        if existing is None:
            seen: set[str] = set()
            item_states: list[dict] = []
            for index, item in enumerate(resolved_items):
                if not isinstance(item, dict):
                    return {
                        "status": "failed",
                        "error": "foreach_subworkflow items must be objects",
                    }
                key = str(item.get(item_key_field) or "").strip()
                if not key:
                    return {
                        "status": "failed",
                        "error": f"foreach_subworkflow item {index} has no {item_key_field}",
                    }
                if key in seen:
                    return {
                        "status": "failed",
                        "error": f"foreach_subworkflow has duplicate item key: {key}",
                    }
                seen.add(key)
                item_states.append({
                    "key": key,
                    "index": index,
                    "item": item,
                    "status": "pending",
                    "subrun_id": None,
                    "attempts": 0,
                    "output": None,
                    "error": None,
                })
            progress = {
                "input_fingerprint": input_fingerprint,
                "concurrency": concurrency,
                "items": item_states,
            }
            self._set_foreach_progress(run, progress_key, progress)
        else:
            progress = dict(existing)
            if progress.get("input_fingerprint") != input_fingerprint:
                return {
                    "status": "failed",
                    "error": "foreach_subworkflow inputs changed while child runs were active",
                }
            item_states = [dict(item) for item in progress.get("items") or []]
            progress["items"] = item_states

        target = (await db.execute(
            select(WorkflowDefinition).where(
                WorkflowDefinition.entity_id == run.entity_id,
                or_(
                    WorkflowDefinition.id == target_id,
                    WorkflowDefinition.name == target_id,
                ),
            ).limit(1)
        )).scalar_one_or_none()
        if target is None:
            return {"status": "failed", "error": f"subworkflow {target_id} not found"}
        if target.id == run.workflow_id:
            return {"status": "failed", "error": "foreach_subworkflow cannot call itself"}

        from packages.core.services.workflow_service import (
            _snapshot_display_metadata,
            entry_step_id,
        )

        child_entry = entry_step_id(target.steps or [])
        if child_entry is None:
            return {
                "status": "failed",
                "error": "subworkflow requires an explicit trigger or webhook entry node",
            }

        retry_raw = config.get("retry_item_keys", [])
        retry_keys_value = _resolve_binding(retry_raw, variables)
        retry_keys = {
            str(value) for value in retry_keys_value
        } if isinstance(retry_keys_value, list) else set()
        max_attempts = max(1, min(int(config.get("max_attempts") or 2), 5))
        for item_state in item_states:
            if (
                item_state.get("status") == "failed"
                and item_state.get("key") in retry_keys
                and int(item_state.get("attempts") or 0) < max_attempts
            ):
                item_state.update({
                    "status": "pending",
                    "subrun_id": None,
                    "output": None,
                    "error": None,
                })

        while True:
            active = [
                item for item in item_states
                if item.get("status") in {"running", "paused"}
            ]
            if active:
                progress["items"] = item_states
                self._set_foreach_progress(run, progress_key, progress)
                return {
                    "status": "paused",
                    "items": self._foreach_public_items(item_states),
                    "subrun_ids": [
                        str(item["subrun_id"])
                        for item in active
                        if item.get("subrun_id")
                    ],
                    "output_var": config.get("output_var"),
                }

            failed = [item for item in item_states if item.get("status") == "failed"]
            if failed:
                progress["items"] = item_states
                self._set_foreach_progress(run, progress_key, progress)
                return {
                    "status": "failed",
                    "error": failed[0].get("error") or "foreach_subworkflow child failed",
                    "items": self._foreach_public_items(item_states),
                    "output_var": config.get("output_var"),
                }

            pending = [item for item in item_states if item.get("status") == "pending"]
            if not pending:
                outputs = [item.get("output") for item in item_states]
                updated = dict(run.variables or {})
                updated.pop(progress_key, None)
                run.variables = updated
                return {
                    "status": "completed",
                    "output": outputs,
                    "items": self._foreach_public_items(item_states),
                    "output_var": config.get("output_var"),
                }

            batch = pending[:concurrency]
            child_runs: list[WorkflowRun] = []
            for item_state in batch:
                child_scope = dict(variables)
                child_scope[item_var] = item_state["item"]
                child_scope["index"] = item_state["index"]
                _bind_inputs(config, child_scope)
                seed = dict(target.variables or {})
                seed[item_var] = item_state["item"]
                seed["index"] = item_state["index"]
                for input_spec in config.get("inputs") or []:
                    if not isinstance(input_spec, dict):
                        continue
                    input_name = str(
                        input_spec.get("key") or input_spec.get("name") or ""
                    ).strip()
                    if input_name and input_name in child_scope:
                        seed[input_name] = child_scope[input_name]

                target_fingerprint = workflow_definition_fingerprint(target)
                target_snapshot = build_definition_snapshot(
                    target,
                    fingerprint=target_fingerprint,
                )
                child_run_id = generate_ulid()
                child_run = WorkflowRun(
                    id=child_run_id,
                    workflow_id=target.id,
                    entity_id=run.entity_id,
                    workspace_id=run.workspace_id,
                    lineage_root_run_id=child_run_id,
                    lineage_is_legacy=False,
                    attempt_number=1,
                    status="running",
                    current_step_id=child_entry,
                    variables=seed,
                    step_results={},
                    definition_snapshot=target_snapshot,
                    execution_trace=[],
                    trigger_data={
                        **_snapshot_display_metadata(target_snapshot),
                        "_subworkflow_depth": depth + 1,
                        "parent_run_id": run.id,
                        "parent_step_id": step.get("id"),
                        "parent_foreach_item_key": item_state["key"],
                        "parent_foreach_progress_key": progress_key,
                        "_workflow_definition_version": target_snapshot["version"],
                        "_workflow_definition_fingerprint": target_fingerprint,
                        "_workflow_runtime_context": _child_workflow_context(
                            run,
                            seed,
                        ),
                    },
                    started_by=run.started_by,
                    started_at=_utc_now(),
                )
                db.add(child_run)
                child_runs.append(child_run)
                item_state["status"] = "running"
                item_state["subrun_id"] = child_run.id
                item_state["attempts"] = int(item_state.get("attempts") or 0) + 1

            progress["items"] = item_states
            self._set_foreach_progress(run, progress_key, progress)
            await db.flush()
            await db.commit()

            await asyncio.gather(*(
                WorkflowRunner(cache_index=self._cache_index).run(child.id)
                for child in child_runs
            ))
            for child in child_runs:
                await db.refresh(child)
                item_state = next(
                    item for item in item_states if item.get("subrun_id") == child.id
                )
                item_state["status"] = child.status
                item_state["output"] = (
                    dict(child.variables or {}) if child.status == "completed" else None
                )
                item_state["error"] = child.error if child.status == "failed" else None
            progress["items"] = item_states
            self._set_foreach_progress(run, progress_key, progress)

    # ── Unsupported step (imported, unmapped) ────────────────────────────

    def _execute_unsupported_step(self, step: dict) -> dict:
        """Gracefully skip a node that had no manor equivalent on import.

        Imported workflows (ComfyUI/n8n/Dify) keep unmapped nodes as
        ``unsupported`` placeholders so the graph stays intact. By default we
        skip them and continue; set ``config.on_unsupported = "fail"`` to make
        an unmapped node halt the run instead.
        """
        meta = step.get("meta", {})
        original_type = meta.get("original_type", "unknown")
        source = meta.get("source_tool", "import")
        msg = f"Skipped unsupported node (from {source}, original type '{original_type}')"
        logger.warning("workflow_runner: %s [step=%s]", msg, step.get("id"))
        if step.get("config", {}).get("on_unsupported") == "fail":
            return {"status": "failed", "error": msg}
        return {"status": "completed", "output": msg, "skipped": True}

    # ── Condition evaluator ──────────────────────────────────────────────

    def _evaluate_condition(self, step: dict, variables: dict) -> bool:
        """Evaluate a condition expression against workflow variables.

        Supports:
          - single comparisons: ``score > 0.7``, ``status == "approved"``,
            ``len(items) > 0``, ``result.success == true``
          - compound boolean logic: ``score > 0.7 and status == "approved"``,
            ``a == 1 or b == 2`` (``or`` lowest precedence, then ``and``)

        Uses operator-based comparison with safe value resolution. No eval().
        """
        config = step.get("config", {})
        expression = str(config.get("expression", "true")).strip()
        return self._eval_bool_expr(expression, variables)

    def _eval_bool_expr(self, expression: str, variables: dict) -> bool:
        """Evaluate a (possibly compound) boolean expression."""
        expression = expression.strip()
        # OR has the lowest precedence
        or_parts = re.split(r"\s+or\s+", expression, flags=re.IGNORECASE)
        if len(or_parts) > 1:
            return any(self._eval_bool_expr(p, variables) for p in or_parts)
        and_parts = re.split(r"\s+and\s+", expression, flags=re.IGNORECASE)
        if len(and_parts) > 1:
            return all(self._eval_bool_expr(p, variables) for p in and_parts)
        return self._eval_atom(expression, variables)

    def _eval_atom(self, expression: str, variables: dict) -> bool:
        """Evaluate a single comparison clause (no and/or)."""
        expression = expression.strip()

        # Handle ``len(var) OP value`` pattern
        len_match = re.match(r"len\((.+?)\)\s*(==|!=|>=|<=|>|<)\s*(.+)", expression)
        if len_match:
            var_name = len_match.group(1)
            op_str = len_match.group(2)
            right_raw = len_match.group(3)
            val = _lookup_reference(var_name, variables)
            left_val = len(val) if hasattr(val, "__len__") else 0
            right_val = _resolve_value(right_raw, variables)
            try:
                right_val = float(right_val)
                left_val = float(left_val)
            except (ValueError, TypeError):
                pass
            op_fn = _OPS.get(op_str, operator.eq)
            return bool(op_fn(left_val, right_val))

        # Standard ``left OP right`` pattern
        for op_str in sorted(_OPS, key=len, reverse=True):
            if op_str in expression:
                parts = expression.split(op_str, 1)
                if len(parts) == 2:
                    left_val = _resolve_value(parts[0], variables)
                    right_val = _resolve_value(parts[1], variables)

                    # Numeric comparison if both sides can be numbers
                    try:
                        left_val = float(left_val)
                        right_val = float(right_val)
                    except (ValueError, TypeError):
                        left_val = str(left_val).strip().strip("'\"")
                        right_val = str(right_val).strip().strip("'\"")

                    op_fn = _OPS[op_str]
                    return bool(op_fn(left_val, right_val))

        # Bare variable name — truthy check
        val = _lookup_reference(expression, variables, missing=False)
        return bool(val)

    # ── Step resolution ──────────────────────────────────────────────────

    def _find_runnable_steps(
        self, workflow_def: WorkflowDefinition, run: WorkflowRun,
    ) -> list[dict]:
        """Find steps that are ready to execute (dependencies met).

        A step is runnable when:
        1. It has not been executed yet (no entry in step_results)
        2. All steps listed in its ``depends_on`` have completed successfully
        3. It is reachable from the current execution path

        For graph workflows linked by ``next``, traversal starts at the entry
        node and follows only edges selected by completed branch nodes. Every
        reachable frontier node is returned, so fan-out and fan-in work without
        losing branches to a single ``current_step_id`` pointer. For legacy DAG
        workflows with ``depends_on``, all dependency-ready steps are returned.
        """
        steps = workflow_def.steps or []
        step_results = run.step_results or {}
        traversed_ids = {
            sid for sid, res in step_results.items()
            if res.get("status") == "completed" or res.get("continued") is True
        }

        # Imported/frozen workflows can use dependencies for linear setup and
        # explicit routes only at gates. Pure DAGs keep the legacy scheduler;
        # hybrid workflows are traversed below with dependency edges filling
        # only the gaps where a node has no explicit outgoing route.
        has_graph_routes = any(self._declared_next(step) for step in steps)
        has_deps = any(s.get("depends_on") for s in steps)

        if has_deps and not has_graph_routes:
            # DAG mode: return all steps whose deps are met and not yet run
            runnable = []
            for step in steps:
                sid = step["id"]
                if sid in step_results:
                    continue
                deps = step.get("depends_on", [])
                if all(d in traversed_ids for d in deps):
                    runnable.append(step)
            return runnable

        step_map = {s["id"]: s for s in steps if s.get("type") != "note"}
        if not step_map:
            return []

        dependency_children: dict[str, list[str]] = {}
        if has_deps:
            for step in steps:
                for dependency in step.get("depends_on") or []:
                    dependency_children.setdefault(str(dependency), []).append(step["id"])

        def outgoing(step: dict, result: dict | None = None) -> list[str]:
            if self._declared_next(step):
                return self._selected_next(step, result) if result is not None else self._declared_next(step)
            return dependency_children.get(step["id"], [])

        from packages.core.services.workflow_service import entry_step_id

        if has_deps:
            entries = [
                step["id"]
                for step in steps
                if step.get("type") != "note" and not (step.get("depends_on") or [])
            ]
        else:
            entry = run.current_step_id if not step_results else entry_step_id(steps)
            if entry not in step_map:
                entry = entry_step_id(steps)
            entries = [entry] if entry else []
        if not entries:
            return []

        reachable: set[str] = set()
        queue = list(entries)
        while queue:
            sid = queue.pop(0)
            if sid in reachable or sid not in step_map:
                continue
            reachable.add(sid)
            if sid in traversed_ids:
                queue.extend(outgoing(step_map[sid], step_results.get(sid)))
            else:
                # Follow every declared route through not-yet-run nodes.  This
                # lets a downstream fan-in see indirect active predecessors and
                # wait for them instead of racing a sibling node in the same
                # scheduling iteration (common in imported n8n Merge graphs).
                queue.extend(outgoing(step_map[sid]))

        # Build active incoming edges. Completed branch nodes contribute only
        # their selected route; unexecuted reachable nodes contribute every
        # possible route so a fan-in waits for all active upstream branches.
        active_incoming: dict[str, set[str]] = {}
        for sid in reachable:
            step = step_map[sid]
            targets = (
                outgoing(step, step_results.get(sid))
                if sid in traversed_ids
                else outgoing(step)
            )
            for target in targets:
                active_incoming.setdefault(target, set()).add(sid)

        runnable = []
        for step in steps:
            sid = step.get("id")
            if step.get("type") == "note" or sid not in reachable or sid in step_results:
                continue
            explicit_deps = {
                str(dep) for dep in (step.get("depends_on") or []) if dep
            }
            if (
                all(pred in traversed_ids for pred in active_incoming.get(sid, set()))
                and all(dep in traversed_ids for dep in explicit_deps)
            ):
                runnable.append(step)
        return runnable

    @staticmethod
    def _declared_next(step: dict) -> list[str]:
        """All routes a not-yet-executed step could activate."""
        targets: list[str] = []
        for key in ("next", "true_next", "false_next"):
            value = step.get(key) or []
            targets.extend(value if isinstance(value, list) else [value])
        if step.get("type") == "switch":
            config = step.get("config") or {}
            for case in config.get("cases") or []:
                value = case.get("next") or [] if isinstance(case, dict) else []
                targets.extend(value if isinstance(value, list) else [value])
            value = config.get("default_next") or []
            targets.extend(value if isinstance(value, list) else [value])
        return list(dict.fromkeys(str(t) for t in targets if t))

    @classmethod
    def _selected_next(cls, step: dict, result: dict | None) -> list[str]:
        """Routes selected by a completed step, preserving an explicit empty branch."""
        if result is not None and "next_override" in result:
            value = result.get("next_override")
            if value in (None, ""):
                return []
            return value if isinstance(value, list) else [value]
        value = step.get("next") or []
        return value if isinstance(value, list) else [value]

    def _all_steps_done(self, workflow_def: WorkflowDefinition, run: WorkflowRun) -> bool:
        """Check whether every node on the selected execution path completed."""
        steps = workflow_def.steps or []
        step_results = run.step_results or {}
        has_graph_routes = any(self._declared_next(step) for step in steps)
        has_deps = any(s.get("depends_on") for s in steps)
        if has_deps and not has_graph_routes:
            executable = [s for s in steps if s.get("type") != "note"]
            return all(s["id"] in step_results for s in executable)

        traversed = {
            sid for sid, result in step_results.items()
            if result.get("status") == "completed" or result.get("continued") is True
        }
        from packages.core.services.workflow_service import entry_step_id

        step_map = {s["id"]: s for s in steps if s.get("type") != "note"}
        dependency_children: dict[str, list[str]] = {}
        if has_deps:
            for step in steps:
                for dependency in step.get("depends_on") or []:
                    dependency_children.setdefault(str(dependency), []).append(step["id"])

        def outgoing(step: dict, result: dict | None = None) -> list[str]:
            if self._declared_next(step):
                return self._selected_next(step, result) if result is not None else self._declared_next(step)
            return dependency_children.get(step["id"], [])

        if has_deps:
            entries = [
                step["id"]
                for step in steps
                if step.get("type") != "note" and not (step.get("depends_on") or [])
            ]
        else:
            entry = entry_step_id(steps)
            entries = [entry] if entry else []
        if not entries:
            return True
        reachable: set[str] = set()
        queue = list(entries)
        while queue:
            sid = queue.pop(0)
            if sid in reachable or sid not in step_map:
                continue
            reachable.add(sid)
            if sid in traversed:
                queue.extend(outgoing(step_map[sid], step_results.get(sid)))
        return bool(reachable) and reachable.issubset(traversed)

    # ── Result recording ─────────────────────────────────────────────────

    def _record_step_result(
        self, step: dict, result: dict, run: WorkflowRun,
    ) -> None:
        """Store step result and update run variables."""
        step_id = step["id"]

        # Update step_results
        step_results = dict(run.step_results or {})
        step_results[step_id] = result
        run.step_results = step_results

        # Keep the pointer on the completed node. Graph traversal derives every
        # selected outgoing edge from its result, including fan-out branches.
        if result.get("status") == "completed" or result.get("continued") is True:
            run.current_step_id = step_id
            # Data flow: every step's output is stored under its own id (the
            # auto variable, e.g. {{r}} / {{r.context}}), plus an optional
            # explicit ``output_var`` alias. Downstream {{templates}} read these.
            output = result.get("output")
            updated_vars = dict(run.variables or {})
            if output is not None:
                updated_vars[step_id] = output
                output_var = result.get("output_var")
                if output_var:
                    updated_vars[output_var] = output
            # Explicit named outputs (config.outputs): expose specific fields of
            # this step's result under chosen names. A blank value defaults to
            # the whole result; an expression (e.g. {{id.field}}) is resolved
            # against the scope that already includes the step's auto var.
            for item in (step.get("config") or {}).get("outputs") or []:
                if not isinstance(item, dict):
                    continue
                key = str(item.get("key") or item.get("name") or "").strip()
                if not key:
                    continue
                raw = item.get("value")
                resolved = output if raw in (None, "") else _resolve_binding(str(raw), updated_vars)
                updated_vars[key] = _coerce_typed(resolved, item.get("type"))
            if updated_vars != (run.variables or {}):
                run.variables = updated_vars
        elif result.get("status") == "paused":
            run.current_step_id = step_id

    # ── Re-enqueue via Celery ────────────────────────────────────────────

    @classmethod
    async def resume(
        cls,
        workflow_run_id: str,
        variables: dict | None = None,
        *,
        entity_id: str | None = None,
        resumed_by: str | None = None,
        execute: bool = True,
    ) -> str:
        """Atomically resume a paused run and continue after its wait node.

        Returns ``"resumed"``, ``"not_found"``, or ``"not_paused"``. Keeping
        the wait node as ``current_step_id`` while changing its recorded result
        to completed lets the normal graph traversal follow every outgoing edge
        (including fan-out) without special resume-only routing logic.
        """
        async with async_session() as db:
            from sqlalchemy import select

            stmt = select(WorkflowRun).where(WorkflowRun.id == workflow_run_id)
            if entity_id:
                stmt = stmt.where(WorkflowRun.entity_id == entity_id)
            result = await db.execute(stmt.with_for_update())
            run = result.scalar_one_or_none()
            if not run:
                return "not_found"
            if run.status != "paused":
                return "not_paused"

            workflow = (await db.execute(
                select(WorkflowDefinition).where(
                    WorkflowDefinition.id == run.workflow_id,
                    WorkflowDefinition.entity_id == run.entity_id,
                )
            )).scalar_one_or_none()
            if workflow and workflow_definition_changed(workflow, run):
                run.status = "failed"
                run.error = DEFINITION_CHANGED_ERROR
                run.completed_at = _utc_now()
                await db.commit()
                return "definition_changed"
            current_step_id = run.current_step_id
            current_step = next(
                (
                    step
                    for step in (workflow.steps if workflow else [])
                    if step.get("id") == current_step_id
                ),
                None,
            )
            current_config = (
                current_step.get("config")
                if isinstance(current_step, dict)
                and isinstance(current_step.get("config"), dict)
                else {}
            )
            stage_wait_context = workflow_stage_wait_context(run, current_step)
            if isinstance(current_step, dict) and current_step.get("type") == "stage":
                if stage_wait_context is None:
                    return "not_paused"
                internal_wait = stage_wait_context[1]
                current_config = (
                    internal_wait.get("config")
                    if isinstance(internal_wait.get("config"), dict)
                    else {}
                )
            is_approval = current_config.get("wait_type", "approval") == "approval"
            approval_metadata: dict[str, Any] = {}
            if is_approval:
                actor_id = str(resumed_by or "").strip()
                if not actor_id or actor_id != str(run.started_by or "").strip():
                    return "invalid_approval"
                response_variable = str(
                    current_config.get("response_variable") or "decision"
                ).strip()
                response_value = (variables or {}).get(response_variable)
                decision_value = (
                    response_value.get("choice")
                    if isinstance(response_value, dict)
                    else response_value
                )
                decision = str(decision_value or "").strip()
                if not decision:
                    return "invalid_approval"
                try:
                    approval_metadata = workflow_approval_decision_metadata(
                        current_config,
                        decision=decision,
                        actor_id=actor_id,
                    )
                except ValueError:
                    return "invalid_approval"

            updated_vars = dict(run.variables or {})
            if variables:
                updated_vars.update(variables)
            run.variables = updated_vars

            if current_step_id and stage_wait_context is not None:
                complete_workflow_stage_wait(
                    run,
                    current_step,
                    stage_wait_context,
                    metadata=approval_metadata,
                )
            elif current_step_id:
                step_results = dict(run.step_results or {})
                previous = dict(step_results.get(current_step_id) or {})
                completed_at = _utc_now().isoformat()
                previous.update({
                    "status": "completed",
                    "resumed": True,
                    "resumed_at": completed_at,
                    "completed_at": completed_at,
                    **approval_metadata,
                })
                step_results[current_step_id] = previous
                run.step_results = step_results
                append_execution_trace(
                    run,
                    node=current_step or {"id": current_step_id},
                    status="completed",
                    result=previous,
                )

            run.status = "running"
            run.error = None
            await db.commit()

        if execute:
            await cls().run(workflow_run_id)
        else:
            cls.enqueue(workflow_run_id)
        return "resumed"

    @staticmethod
    def enqueue(workflow_run_id: str, delay_seconds: float = 0) -> bool:
        """Dispatch a workflow run to Celery for async execution."""
        try:
            from packages.core.tasks.ai_tasks import run_workflow
            kwargs: dict = {}
            if delay_seconds > 0:
                kwargs["countdown"] = delay_seconds
            run_workflow.apply_async(args=[workflow_run_id], **kwargs)
            return True
        except Exception as exc:
            logger.warning(
                "WorkflowRunner: failed to enqueue run %s: %s",
                workflow_run_id, exc,
            )
            return False

    @staticmethod
    def enqueue_resume(workflow_run_id: str, delay_seconds: float) -> bool:
        """Schedule a paused timer run to resume after ``delay_seconds``."""
        try:
            from packages.core.tasks.ai_tasks import resume_workflow

            resume_workflow.apply_async(
                args=[workflow_run_id],
                countdown=max(float(delay_seconds), 0.0),
            )
            return True
        except Exception as exc:
            logger.warning(
                "WorkflowRunner: failed to schedule resume for run %s: %s",
                workflow_run_id,
                exc,
            )
            return False
