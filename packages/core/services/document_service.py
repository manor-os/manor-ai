"""Document service — CRUD, groups, file metadata."""
from __future__ import annotations

import asyncio
import json
import os
from html.parser import HTMLParser
from typing import Optional

from sqlalchemy import String, and_, func, not_, or_, select, update
from sqlalchemy.ext.asyncio import AsyncSession

from packages.core.models.base import generate_ulid
from packages.core.models.document import Document, DocumentGroup, DocumentGroupMember, VectorStatus
from packages.core.services.document_metadata import merge_document_metadata
from packages.core.services.entity_fs import SYSTEM_DIRS, SYSTEM_FILES
from packages.core.services.knowledge_visibility import HIDDEN_PREFIXES, is_user_visible_path, normalize_rel_path
from packages.core.services.tool_cache_version import bump_tool_cache_version


class StorageLimitExceeded(Exception):
    """Raised when adding a document would exceed the entity's plan storage.

    Carries the same fields as the plan gate's 402 detail so the HTTP layer and
    runtime tools can surface a consistent "upgrade to add more" message.
    """

    def __init__(self, message: str, *, plan: str = "", limit: float | None = None, current: float | None = None):
        super().__init__(message)
        self.message = message
        self.plan = plan
        self.limit = limit
        self.current = current


async def _enforce_storage_limit(db: AsyncSession, entity_id: str) -> None:
    """Raise :class:`StorageLimitExceeded` if the entity is at/over its plan's
    knowledge-base storage limit. No-op in OSS/self-hosted (plan gate allows)."""
    from packages.core.services.plan_gate import check

    gate = await check(db, entity_id, "storage_mb")
    if not gate.allowed:
        raise StorageLimitExceeded(
            gate.message or "Knowledge base storage limit reached. Upgrade for more.",
            plan=gate.plan, limit=gate.limit, current=gate.current,
        )


# ── Documents ──

def _visible_fs_path_clause():
    hidden_checks = [
        Document.fs_path.ilike(".%"),
        Document.fs_path.ilike("%/.%"),
    ]
    for prefix in HIDDEN_PREFIXES:
        bare = prefix.rstrip("/")
        hidden_checks.append(Document.fs_path == bare)
        hidden_checks.append(Document.fs_path.ilike(f"{prefix}%"))
    for name in SYSTEM_FILES:
        hidden_checks.append(Document.fs_path == name)
        hidden_checks.append(Document.fs_path.ilike(f"%/{name}"))
    for dirname in SYSTEM_DIRS:
        hidden_checks.append(Document.fs_path == dirname)
        hidden_checks.append(Document.fs_path.ilike(f"{dirname}/%"))
        hidden_checks.append(Document.fs_path.ilike(f"%/{dirname}/%"))
    return or_(Document.fs_path.is_(None), not_(or_(*hidden_checks)))

async def list_documents(
    db: AsyncSession, entity_id: str, *,
    name_search: str | None = None,
    folder_id: str | None = None,
    folder_ids: set[str] | None = None,
    workspace_id: str | None = None,
    include_generated_assets: bool = True,
    limit: int | None = 100, offset: int = 0,
) -> tuple[list[Document], int]:
    conditions = _document_scope_conditions(
        entity_id,
        name_search=name_search,
        workspace_id=workspace_id,
        include_generated_assets=include_generated_assets,
    )
    q = select(Document).where(*conditions)
    count_q = select(func.count()).select_from(Document).where(*conditions)
    if folder_ids is not None:
        if not folder_ids:
            return [], 0
        q = q.where(Document.folder_id.in_(folder_ids))
        count_q = count_q.where(Document.folder_id.in_(folder_ids))
    elif folder_id is not None:
        # "root" or empty string means documents with no folder
        if folder_id == "" or folder_id == "root":
            q = q.where(Document.folder_id.is_(None))
            count_q = count_q.where(Document.folder_id.is_(None))
        else:
            q = q.where(Document.folder_id == folder_id)
            count_q = count_q.where(Document.folder_id == folder_id)
    q = q.order_by(Document.created_at.desc())
    if limit is not None:
        q = q.limit(limit).offset(offset)
    elif offset:
        q = q.offset(offset)
    result = await db.execute(q)
    count_result = await db.execute(count_q)
    return list(result.scalars().all()), count_result.scalar_one()


def _document_scope_conditions(
    entity_id: str, *,
    name_search: str | None = None,
    workspace_id: str | None = None,
    include_generated_assets: bool = True,
) -> list:
    """Shared WHERE clauses for the visible-document scope.

    Everything that defines *which* documents are in view — entity, not
    trashed, not hidden, generated-asset and workspace filters, name search —
    minus the per-folder filter and pagination. Used by both the paginated
    listing and the storage-usage aggregate so the two always agree.
    """
    conditions = [
        Document.entity_id == entity_id,
        Document.is_trashed == False,  # noqa: E712
        _visible_fs_path_clause(),
    ]
    if not include_generated_assets:
        generated_media = and_(
            Document.source.in_(("ai_generated", "sandbox", "bash", "agent", "elevenlabs", "mcp")),
            or_(
                Document.mime_type.ilike("image/%"),
                Document.mime_type.ilike("video/%"),
                Document.mime_type.ilike("audio/%"),
                Document.file_type.in_(("png", "jpg", "jpeg", "webp", "gif", "mp4", "mov", "webm", "mp3", "wav", "m4a")),
            ),
        )
        conditions.append(not_(generated_media))
    if name_search:
        search = f"%{name_search}%"
        conditions.append(or_(
            Document.name.ilike(search),
            Document.fs_path.ilike(search),
            Document.file_type.ilike(search),
            Document.mime_type.ilike(search),
            Document.source.ilike(search),
            Document.metadata_.cast(String).ilike(search),
        ))
    if workspace_id:
        group_membership = (
            select(DocumentGroupMember.document_id)
            .join(DocumentGroup, DocumentGroup.id == DocumentGroupMember.group_id)
            .where(
                DocumentGroupMember.document_id == Document.id,
                DocumentGroup.entity_id == entity_id,
                DocumentGroup.workspace_id == workspace_id,
            )
            .exists()
        )
        conditions.append(or_(
            group_membership,
            Document.metadata_["origin"]["workspace_id"].astext == workspace_id,
        ))
    return conditions


async def storage_usage(
    db: AsyncSession, entity_id: str, *,
    name_search: str | None = None,
    folder_ids: set[str] | None = None,
    workspace_id: str | None = None,
    include_generated_assets: bool = True,
) -> tuple[int, int]:
    """Total ``(size_bytes, file_count)`` for visible documents in scope.

    Unlike :func:`list_documents`, this is **not** capped to a single folder
    level or page: when ``folder_ids`` is given it sums every document in that
    set of folders (the caller passes a folder plus its descendants), and when
    it is ``None`` it covers the whole scope (the entire knowledge base, or a
    workspace). This is what the Knowledge Base header should show — the size of
    everything under the current location, including nested folders.
    """
    conditions = _document_scope_conditions(
        entity_id,
        name_search=name_search,
        workspace_id=workspace_id,
        include_generated_assets=include_generated_assets,
    )
    q = select(
        func.coalesce(func.sum(Document.file_size), 0),
        func.count(),
    ).select_from(Document).where(*conditions)
    if folder_ids is not None:
        if not folder_ids:
            return 0, 0
        q = q.where(Document.folder_id.in_(folder_ids))
    row = (await db.execute(q)).one()
    return int(row[0] or 0), int(row[1] or 0)


async def get_document(db: AsyncSession, doc_id: str, entity_id: str) -> Optional[Document]:
    result = await db.execute(
        select(Document).where(
            Document.id == doc_id,
            Document.entity_id == entity_id,
            Document.is_trashed == False,  # noqa: E712
            _visible_fs_path_clause(),
        )
    )
    return result.scalar_one_or_none()


async def create_document(
    db: AsyncSession, entity_id: str, *,
    name: str, fs_path: str | None = None, file_url: str | None = None,
    file_size: int | None = None, file_type: str | None = None,
    mime_type: str | None = None, source: str = "upload",
    created_by: str | None = None, folder_id: str | None = None,
    metadata: dict | None = None,
    # ── Permission-v1 fields (see docs/PERMISSIONS_DESIGN_ZH.md §13) ────
    visibility: str | None = None,
    classification: str | None = None,
    client_visible: bool | None = None,
    owner_id: str | None = None,
    # When True, skip the plan storage-limit check. Used for bookkeeping that
    # re-projects files already on disk (e.g. filesystem reconcile), which must
    # not be blocked just because the entity is over its quota.
    skip_storage_check: bool = False,
) -> Document:
    # Every new knowledge-base document funnels through here (direct creates and
    # the new-row branch of upsert_document_by_fs_path), so this is the single
    # chokepoint that enforces the storage limit across ALL add paths — uploads,
    # AI drafts, URL/Drive imports, and agent/sandbox-generated files alike.
    if not skip_storage_check:
        await _enforce_storage_limit(db, entity_id)
    doc = Document(
        id=generate_ulid(),
        entity_id=entity_id,
        name=name,
        fs_path=fs_path,
        file_url=file_url,
        file_size=file_size,
        file_type=file_type,
        mime_type=mime_type,
        source=source,
        created_by=created_by,
        folder_id=folder_id,
    )
    if visibility is None and owner_id is not None and folder_id is None:
        doc.visibility = "private"
    if visibility is not None:
        doc.visibility = visibility
    if classification is not None:
        doc.classification = classification
    if client_visible is not None:
        doc.client_visible = client_visible
    if owner_id is not None:
        doc.owner_id = owner_id
    if metadata:
        doc.metadata_ = merge_document_metadata(metadata)
    db.add(doc)
    await db.flush()

    from packages.core.services.event_emitter import emit
    emit(entity_id, "document.uploaded", source="document_service", payload={"document_id": doc.id, "name": name})
    await bump_tool_cache_version(entity_id, "documents")

    return doc


async def upsert_document_by_fs_path(
    db: AsyncSession,
    entity_id: str,
    *,
    fs_path: str,
    name: str,
    file_size: int | None = None,
    file_type: str | None = None,
    mime_type: str | None = None,
    source: str = "manual",
    created_by: str | None = None,
    folder_id: str | None = None,
    skip_storage_check: bool = False,
) -> Document:
    """Idempotent upsert by (entity_id, fs_path).

    Updating an existing projection never counts as "adding" storage, so the
    limit is only enforced on the new-row branch (and skippable for reconcile).
    """
    existing = await db.execute(
        select(Document).where(
            Document.entity_id == entity_id,
            Document.fs_path == fs_path,
        ).limit(1)
    )
    doc = existing.scalar_one_or_none()
    if doc:
        doc.name = name
        doc.file_size = file_size
        doc.file_type = file_type
        doc.mime_type = mime_type
        # A filesystem write can recreate a path that was previously soft
        # deleted. In that case the filesystem is the source of truth again, so
        # revive the Knowledge projection instead of leaving it hidden in Trash.
        doc.is_trashed = False
        doc.trashed_at = None
        doc.trashed_by = None
        if folder_id is not None:
            doc.folder_id = folder_id
        if not doc.source:
            doc.source = source
        if created_by and not doc.created_by:
            doc.created_by = created_by
        await bump_tool_cache_version(entity_id, "documents")
        return doc
    return await create_document(
        db, entity_id,
        name=name,
        fs_path=fs_path,
        file_size=file_size,
        file_type=file_type,
        mime_type=mime_type,
        source=source,
        created_by=created_by,
        folder_id=folder_id,
        skip_storage_check=skip_storage_check,
    )


async def rename_document(db: AsyncSession, doc_id: str, entity_id: str, new_name: str) -> Optional[Document]:
    """Rename a document. Returns the updated document or None if not found."""
    doc = await get_document(db, doc_id, entity_id)
    if not doc:
        return None
    doc.name = new_name
    await db.flush()
    await bump_tool_cache_version(entity_id, "documents")
    return doc


async def delete_document(db: AsyncSession, doc_id: str, entity_id: str) -> bool:
    doc = await get_document(db, doc_id, entity_id)
    if not doc:
        return False
    await db.delete(doc)
    await db.flush()
    await bump_tool_cache_version(entity_id, "documents")
    return True


# ── Document Groups ──

async def list_groups(db: AsyncSession, entity_id: str) -> list[DocumentGroup]:
    result = await db.execute(
        select(DocumentGroup).where(DocumentGroup.entity_id == entity_id)
    )
    return list(result.scalars().all())


async def create_group(db: AsyncSession, entity_id: str, *, name: str, workspace_id: str | None = None) -> DocumentGroup:
    group = DocumentGroup(
        id=generate_ulid(), entity_id=entity_id,
        name=name, workspace_id=workspace_id,
    )
    db.add(group)
    await db.flush()
    return group


async def trigger_reindex(db: AsyncSession, entity_id: str) -> int:
    """Reset all documents to pending and trigger re-indexing."""
    result = await db.execute(
        update(Document)
        .where(
            Document.entity_id == entity_id,
            Document.vector_status != VectorStatus.PENDING,
            Document.is_trashed == False,  # noqa: E712
            _visible_fs_path_clause(),
        )
        .values(vector_status=VectorStatus.PENDING)
    )
    await db.flush()
    if result.rowcount:
        await bump_tool_cache_version(entity_id, "documents")
    return result.rowcount


async def add_document_to_group(
    db: AsyncSession,
    doc_id: str,
    group_id: str,
    *,
    entity_id: str | None = None,
) -> bool:
    if entity_id:
        doc = await get_document(db, doc_id, entity_id)
        if not doc:
            return False
        group = (await db.execute(
            select(DocumentGroup).where(
                DocumentGroup.id == group_id,
                DocumentGroup.entity_id == entity_id,
            )
        )).scalar_one_or_none()
        if not group:
            return False
    existing = await db.execute(
        select(DocumentGroupMember).where(
            DocumentGroupMember.document_id == doc_id,
            DocumentGroupMember.group_id == group_id,
        )
    )
    if existing.scalar_one_or_none():
        return False
    db.add(DocumentGroupMember(document_id=doc_id, group_id=group_id))
    await db.flush()
    return True


# ── Content read/write ──

def _metadata_text_content(doc: Document) -> str | None:
    """Return legacy inline document text stored before FS projection existed."""
    meta = doc.metadata_ if isinstance(doc.metadata_, dict) else {}
    for key in ("content", "content_text"):
        value = meta.get(key)
        if isinstance(value, str):
            return value
    return None


class _EditorHtmlToDocgenText(HTMLParser):
    """Convert lightweight contentEditable HTML into docgen markdown-ish text."""

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.parts: list[str] = []
        self._inline_marks: list[str] = []
        self._skip_depth = 0

    def _newline(self) -> None:
        if self.parts and not self.parts[-1].endswith("\n"):
            self.parts.append("\n")

    def handle_starttag(self, tag: str, attrs) -> None:  # noqa: ANN001
        tag = tag.lower()
        if self._skip_depth:
            self._skip_depth += 1
            return
        attrs_dict = {str(k).lower(): str(v) for k, v in attrs}
        if attrs_dict.get("data-docx-page-break") == "true":
            self._newline()
            self.parts.append("[[PAGE_BREAK]]")
            self._newline()
            self._skip_depth = 1
            return
        if tag in {"p", "div", "section", "article", "blockquote"}:
            self._newline()
        elif tag in {"h1", "h2", "h3", "h4", "h5", "h6"}:
            self._newline()
            level = int(tag[1])
            self.parts.append("#" * level + " ")
        elif tag == "br":
            self.parts.append("\n")
        elif tag == "li":
            self._newline()
            self.parts.append("- ")
        elif tag in {"strong", "b"}:
            self.parts.append("**")
            self._inline_marks.append("**")
        elif tag in {"em", "i"}:
            self.parts.append("*")
            self._inline_marks.append("*")
        elif tag == "u":
            self.parts.append("[[U]]")
            self._inline_marks.append("[[/U]]")
        elif tag in {"s", "strike", "del"}:
            self.parts.append("[[S]]")
            self._inline_marks.append("[[/S]]")
        elif tag == "span":
            style = attrs_dict.get("style", "").lower()
            closers: list[str] = []
            if "font-weight" in style and ("bold" in style or "700" in style or "800" in style or "900" in style):
                self.parts.append("**")
                closers.append("**")
            if "font-style" in style and "italic" in style:
                self.parts.append("*")
                closers.append("*")
            if "text-decoration" in style and "underline" in style:
                self.parts.append("[[U]]")
                closers.append("[[/U]]")
            if "text-decoration" in style and ("line-through" in style or "strike" in style):
                self.parts.append("[[S]]")
                closers.append("[[/S]]")
            self._inline_marks.append("".join(reversed(closers)))
        elif tag in {"td", "th"}:
            if self.parts and not self.parts[-1].endswith(("| ", "\n")):
                self.parts.append(" | ")

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if self._skip_depth:
            self._skip_depth -= 1
            return
        if tag in {"strong", "b", "em", "i", "u", "s", "strike", "del", "span"} and self._inline_marks:
            closing = self._inline_marks.pop()
            if closing:
                self.parts.append(closing)
        elif tag in {"p", "div", "section", "article", "blockquote", "li"}:
            self._newline()
        elif tag in {"h1", "h2", "h3", "h4", "h5", "h6", "tr"}:
            self._newline()
        elif tag in {"td", "th"}:
            self.parts.append(" | ")

    def handle_data(self, data: str) -> None:
        if self._skip_depth:
            return
        if data:
            self.parts.append(data)

    def text(self) -> str:
        raw = "".join(self.parts)
        lines = []
        for line in raw.splitlines():
            cleaned = " ".join(line.split())
            if cleaned:
                lines.append(cleaned)
        return "\n".join(lines).strip()


def _editor_html_to_docgen_text(content: str) -> str:
    if "<" not in content or ">" not in content:
        return content
    parser = _EditorHtmlToDocgenText()
    parser.feed(content)
    parser.close()
    return parser.text() or content


SPREADSHEET_EDITOR_PAYLOAD_PREFIX = "__MANOR_SPREADSHEET_EDITOR_V1__\n"
PPTX_EDITOR_PAYLOAD_PREFIX = "__MANOR_PPTX_EDITOR_V1__\n"
SPREADSHEET_CHARTS_SHEET = "_manor_charts"


def _parse_spreadsheet_editor_payload(content: str) -> tuple[list[list[object]], list[dict], dict] | None:
    if not content.startswith(SPREADSHEET_EDITOR_PAYLOAD_PREFIX):
        return None
    try:
        payload = json.loads(content[len(SPREADSHEET_EDITOR_PAYLOAD_PREFIX):])
    except json.JSONDecodeError:
        return None
    data = payload.get("data") if isinstance(payload, dict) else None
    charts = payload.get("charts") if isinstance(payload, dict) else None
    styles = payload.get("styles") if isinstance(payload, dict) else None
    if not isinstance(data, list):
        return None
    normalized_data = [
        row if isinstance(row, list) else [row]
        for row in data
    ] or [[""]]
    normalized_charts = charts if isinstance(charts, list) else []
    normalized_styles = styles if isinstance(styles, dict) else {}
    return normalized_data, [chart for chart in normalized_charts if isinstance(chart, dict)], normalized_styles


def _coerce_spreadsheet_cell(value: object, *, force_numeric: bool = False) -> object:
    if value is None or isinstance(value, (int, float, bool)):
        return value
    text = str(value)
    stripped = text.strip()
    if stripped.startswith("="):
        return stripped
    if not stripped:
        return ""
    if not force_numeric:
        return text
    normalized = stripped
    for symbol in "$¥€£₹₩₽₺₫₴₪₦₱฿₡₲₵₭₮₸₼₾₿":
        normalized = normalized.replace(symbol, "")
    normalized = normalized.replace(",", "").replace(" ", "")
    if normalized.endswith("%"):
        normalized = normalized[:-1]
    try:
        if "." in normalized:
            return float(normalized)
        return int(normalized)
    except ValueError:
        return text


def _spreadsheet_content_to_xlsx_bytes(content: str) -> bytes:
    """Convert editor spreadsheet content into a real XLSX workbook."""
    import csv
    import io

    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    parsed_payload = _parse_spreadsheet_editor_payload(content)
    if parsed_payload:
        rows, charts, styles = parsed_payload
    else:
        rows = list(csv.reader(io.StringIO(content))) or [[""]]
        charts = []
        styles = {}

    wb = Workbook()
    ws = wb.active
    row_count = max(1, len(rows))
    col_count = max(1, *(len(row) for row in rows))
    numeric_cells: set[tuple[int, int]] = set()
    for idx, chart_config in enumerate(charts):
        try:
            chart_type = chart_config.get("type")
            label_col = min(max(int(chart_config.get("labelColumn", 0)) + 1, 1), col_count)
            value_col = min(max(int(chart_config.get("valueColumn", 1)) + 1, 1), col_count)
            start_row = min(max(int(chart_config.get("startRow", 1)) + 1, 1), row_count)
            end_row = min(max(int(chart_config.get("endRow", row_count - 1)) + 1, start_row), row_count)
            title = str(chart_config.get("title") or "Chart")
        except (TypeError, ValueError):
            continue

        for row_idx in range(start_row, end_row + 1):
            numeric_cells.add((row_idx, value_col))

    for row_idx, row in enumerate(rows, start=1):
        ws.append([
            _coerce_spreadsheet_cell(value, force_numeric=(row_idx, col_idx) in numeric_cells)
            for col_idx, value in enumerate(row, start=1)
        ])

    def _hex_color(value: object, fallback: str | None = None) -> str | None:
        if not isinstance(value, str):
            return fallback
        stripped = value.strip().lstrip("#")
        if len(stripped) == 6 and all(ch in "0123456789abcdefABCDEF" for ch in stripped):
            return stripped.upper()
        return fallback

    if isinstance(styles, dict):
        for key, raw_style in styles.items():
            if not isinstance(key, str) or ":" not in key or not isinstance(raw_style, dict):
                continue
            try:
                row_idx, col_idx = [int(part) + 1 for part in key.split(":", 1)]
            except ValueError:
                continue
            if row_idx < 1 or col_idx < 1:
                continue
            cell = ws.cell(row=row_idx, column=col_idx)
            font_color = _hex_color(raw_style.get("color"))
            font_size = raw_style.get("fontSize")
            try:
                font_size_value = max(8, min(72, int(font_size))) if font_size is not None else None
            except (TypeError, ValueError):
                font_size_value = None
            cell.font = Font(
                name=str(raw_style.get("fontFamily") or "Calibri"),
                bold=bool(raw_style.get("bold")),
                italic=bool(raw_style.get("italic")),
                size=font_size_value or 11,
                color=font_color,
            )
            fill_color = _hex_color(raw_style.get("fill"))
            if fill_color:
                cell.fill = PatternFill("solid", fgColor=fill_color)
            align = raw_style.get("align")
            if align in {"left", "center", "right"}:
                cell.alignment = Alignment(horizontal=align)

    for idx, chart_config in enumerate(charts):
        try:
            chart_type = chart_config.get("type")
            label_col = min(max(int(chart_config.get("labelColumn", 0)) + 1, 1), col_count)
            value_col = min(max(int(chart_config.get("valueColumn", 1)) + 1, 1), col_count)
            start_row = min(max(int(chart_config.get("startRow", 1)) + 1, 1), row_count)
            end_row = min(max(int(chart_config.get("endRow", row_count - 1)) + 1, start_row), row_count)
            title = str(chart_config.get("title") or "Chart")
        except (TypeError, ValueError):
            continue

        if chart_type == "line":
            chart = LineChart()
        elif chart_type == "pie":
            chart = PieChart()
        else:
            chart = BarChart()

        chart.title = title
        chart.height = 7
        chart.width = 12
        data_ref = Reference(ws, min_col=value_col, max_col=value_col, min_row=start_row, max_row=end_row)
        categories_ref = Reference(ws, min_col=label_col, max_col=label_col, min_row=start_row, max_row=end_row)
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(categories_ref)
        ws.add_chart(chart, f"{get_column_letter(col_count + 2)}{2 + idx * 16}")

    if charts or styles:
        meta = wb.create_sheet(SPREADSHEET_CHARTS_SHEET)
        meta["A1"] = json.dumps({"charts": charts, "styles": styles}, ensure_ascii=False)
        meta.sheet_state = "hidden"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _allocate_document_fs_path(doc: Document, entity_root: str) -> str:
    """Allocate a visible relative path for a legacy metadata-only document."""
    raw_name = normalize_rel_path(doc.name or "")
    filename = os.path.basename(raw_name) or f"document-{doc.id}.txt"
    if "." not in filename and doc.file_type:
        filename = f"{filename}.{str(doc.file_type).lstrip('.')}"
    if not is_user_visible_path(filename):
        ext = f".{str(doc.file_type).lstrip('.')}" if doc.file_type else ".txt"
        filename = f"document-{doc.id}{ext}"

    os.makedirs(entity_root, exist_ok=True)
    root_norm = os.path.normpath(entity_root)
    target = os.path.normpath(os.path.join(entity_root, filename))
    if os.path.commonpath([root_norm, target]) != root_norm:
        raise ValueError("Document path escaped entity root")
    if os.path.exists(target):
        base, ext = os.path.splitext(filename)
        target = os.path.join(entity_root, f"{base}_{doc.id[:8]}{ext}")
    return os.path.relpath(target, entity_root)

async def get_document_content(db: AsyncSession, document_id: str, entity_id: str) -> str | None:
    """Read document content from filesystem."""
    doc = await get_document(db, document_id, entity_id)
    if not doc:
        return None
    if not doc.fs_path:
        return _metadata_text_content(doc)
    from packages.core.config import get_settings
    fs_root = get_settings().MANOR_FS_ROOT
    full_path = os.path.join(fs_root, doc.entity_id, doc.fs_path)

    def _read():
        if not os.path.isfile(full_path):
            return None
        with open(full_path, "r", encoding="utf-8", errors="replace") as f:
            return f.read()

    content = await asyncio.to_thread(_read)
    if content is not None:
        return content
    return _metadata_text_content(doc)


async def save_document_content(
    db: AsyncSession, document_id: str, entity_id: str, content: str,
    *, created_by: str | None = None,
) -> bool:
    """Save document content to filesystem and create a version.

    For presentation files (.pptx), the text content is converted to a real
    binary PPTX via ``docgen_service.generate_pptx`` so the viewer and
    download always serve a proper PowerPoint file.
    """
    doc = await get_document(db, document_id, entity_id)
    if not doc:
        return False
    from packages.core.config import get_settings
    settings = get_settings()
    if not doc.fs_path and not settings.MANOR_FS_ENABLED:
        meta = dict(doc.metadata_ or {})
        meta["content"] = content
        doc.metadata_ = meta
        doc.file_size = len(content.encode("utf-8"))
        await db.flush()
        await bump_tool_cache_version(entity_id, "documents")
        from packages.core.services.version_service import create_version
        await create_version(db, document_id, entity_id, change_summary="Edited", created_by=created_by)
        return True
    fs_root = settings.MANOR_FS_ROOT
    entity_root = os.path.join(fs_root, doc.entity_id)
    if not doc.fs_path:
        doc.fs_path = _allocate_document_fs_path(doc, entity_root)

    ext = os.path.splitext(doc.name or "")[1].lower()
    persisted_bytes: bytes
    if ext in (".docx", ".doc"):
        from packages.core.services.docgen_service import generate_docx

        title = os.path.splitext(doc.name or "Document")[0] or "Document"
        persisted_bytes = await generate_docx(title, _editor_html_to_docgen_text(content))
        doc.file_size = len(persisted_bytes)
        doc.file_type = "docx"
        doc.mime_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    elif ext in (".xlsx", ".xls"):
        persisted_bytes = await asyncio.to_thread(_spreadsheet_content_to_xlsx_bytes, content)
        doc.file_size = len(persisted_bytes)
        doc.file_type = "xlsx"
        doc.mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif ext in (".pptx", ".ppt"):
        persisted_bytes = await _content_to_pptx(doc.name, content)
        doc.file_size = len(persisted_bytes)
        doc.mime_type = "application/vnd.openxmlformats-officedocument.presentationml.presentation"
    else:
        persisted_bytes = content.encode("utf-8")
        doc.file_size = len(persisted_bytes)

    from packages.core.services.entity_fs import write_entity_file_atomic

    await asyncio.to_thread(
        write_entity_file_atomic,
        doc.entity_id,
        doc.fs_path,
        persisted_bytes,
        expected_size=len(persisted_bytes),
        allow_empty=True,
    )

    await db.flush()
    await bump_tool_cache_version(entity_id, "documents")
    # Create version
    from packages.core.services.version_service import create_version
    await create_version(db, document_id, entity_id, change_summary="Edited", created_by=created_by)
    return True


async def save_document_file(
    db: AsyncSession,
    document_id: str,
    entity_id: str,
    file_bytes: bytes,
    *,
    filename: str | None = None,
    mime_type: str | None = None,
    created_by: str | None = None,
) -> Document | None:
    """Replace a document's binary file content on disk."""
    doc = await get_document(db, document_id, entity_id)
    if not doc:
        return None

    from packages.core.config import get_settings
    settings = get_settings()
    if not settings.MANOR_FS_ENABLED:
        raise ValueError("Filesystem storage is not enabled")

    entity_root = os.path.join(settings.MANOR_FS_ROOT, doc.entity_id)
    if not doc.fs_path:
        if filename:
            raw_name = normalize_rel_path(filename)
            visible_name = os.path.basename(raw_name)
            if visible_name and is_user_visible_path(visible_name):
                doc.name = visible_name
        doc.fs_path = _allocate_document_fs_path(doc, entity_root)

    full_path = os.path.realpath(os.path.join(entity_root, doc.fs_path))
    root = os.path.realpath(entity_root)
    if os.path.commonpath([root, full_path]) != root:
        raise ValueError("Document path escaped entity root")

    from packages.core.services.entity_fs import write_entity_file_atomic

    await asyncio.to_thread(
        write_entity_file_atomic,
        doc.entity_id,
        doc.fs_path,
        file_bytes,
        expected_size=len(file_bytes),
        allow_empty=True,
    )

    ext = os.path.splitext(doc.name or filename or "")[1].lstrip(".").lower()
    doc.file_size = len(file_bytes)
    doc.file_type = ext or doc.file_type
    doc.mime_type = mime_type or doc.mime_type or "application/octet-stream"
    doc.vector_status = VectorStatus.PENDING

    await db.flush()
    await bump_tool_cache_version(entity_id, "documents")

    from packages.core.services.version_service import create_version
    await create_version(db, document_id, entity_id, change_summary="Edited file", created_by=created_by)
    return doc


async def _content_to_pptx(doc_name: str, content: str) -> bytes:
    """Convert editor text content to binary PPTX via docgen_service."""
    import re
    from packages.core.services.docgen_service import generate_pptx

    # Extract title from first heading or document name
    title_match = re.match(r"^#\s+(.+)", content, re.MULTILINE)
    title = title_match.group(1).strip() if title_match else os.path.splitext(doc_name)[0]

    return await generate_pptx(title, content)
